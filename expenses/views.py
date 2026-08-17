import csv
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import authenticate
from django.db.models import Sum, Q, F, Window
from django.db.models.functions import Coalesce, TruncMonth
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.authtoken.models import Token
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import AllowAny, IsAuthenticated, BasePermission
from rest_framework.response import Response

from rest_framework.pagination import PageNumberPagination

from .models import (
    Branch, Expense, PaymentModeBalance, BillingReminder, PettyCashDebit, Invoice, DeliveryChallan, PurchaseBill, PurchaseOrder, PaymentReceipt, Quote, BillOfSupply, TaxInvoice, BankStatementEntry, EngineerPnl, SleekBillInvoice, Subscription, AppSetting,
    UserProfile, ALL_SECTIONS, SECTION_DASHBOARD, SECTION_EXPENSES, SECTION_PNL, SECTION_REGION, SECTION_INVOICE, SECTION_CHALLAN, SECTION_PURCHASE, SECTION_PORDER, SECTION_RECEIPT, SECTION_PETTYCASH, SECTION_QUOTE, SECTION_BOS, SECTION_TAXINVOICE, SECTION_IDFC, SECTION_BOB, SECTION_ENGPNL, SECTION_SBINVOICE, SECTION_SUBSCRIPTION, SECTION_INSIGHTS,
)
from .serializers import BranchSerializer, ExpenseSerializer, ExpenseCreateSerializer, PaymentModeBalanceSerializer, BillingReminderSerializer, PettyCashDebitSerializer, InvoiceSerializer, DeliveryChallanSerializer, PurchaseBillSerializer, PurchaseOrderSerializer, PaymentReceiptSerializer, QuoteSerializer, BillOfSupplySerializer, TaxInvoiceSerializer, BankStatementEntrySerializer, EngineerPnlSerializer, SleekBillInvoiceSerializer, SubscriptionSerializer


# ---------------------------------------------------------------------------
# Section-based access control
# ---------------------------------------------------------------------------
# Each non-admin user has a UserProfile.allowed_sections list controlling which
# app pages they may reach. Access is enforced server-side so it can't be
# bypassed by calling the API directly. Admins (is_staff / is_superuser) get
# every section implicitly. Endpoints declare which section they belong to via
# RequireSection(...).

# Legacy group used before per-user sections existed. Users still in it (with no
# profile) are treated as Dashboard + P&L, preserving old behaviour.
PNL_ONLY_GROUP = 'pnl_only'


def is_admin_user(user):
    """Admins see everything and can manage other users."""
    return bool(user and user.is_authenticated and (user.is_staff or user.is_superuser))


def get_allowed_sections(user):
    """The list of section keys this user may access (canonical order)."""
    if not user or not user.is_authenticated:
        return []
    if is_admin_user(user):
        return list(ALL_SECTIONS)
    # Per-user profile is the source of truth.
    profile = getattr(user, 'profile', None)
    if profile is not None:
        sections = profile.clean_sections()
        if sections:
            return sections
        # An empty profile list means "no restriction configured yet" only if
        # they're also not in the legacy group; otherwise fall through.
    # Legacy pnl_only group → Dashboard + P&L.
    if user.groups.filter(name=PNL_ONLY_GROUP).exists():
        return [SECTION_DASHBOARD, SECTION_PNL]
    # No profile, no group → full access (ordinary account).
    return list(ALL_SECTIONS)


def has_section(user, section):
    return section in get_allowed_sections(user)


def is_pnl_only(user):
    """Back-compat flag: user can see P&L but NOT the raw Expenses ledger."""
    if not user or not user.is_authenticated or is_admin_user(user):
        return False
    sections = get_allowed_sections(user)
    return SECTION_PNL in sections and SECTION_EXPENSES not in sections


class RequireSection(BasePermission):
    """Permission factory: only allow users whose sections include `section`.

    Usage:  permission_classes = [IsAuthenticated, RequireSection(SECTION_EXPENSES)]
    or as a decorator arg on function views.
    """
    section = None
    message = 'Your account does not have access to this section.'

    def __init__(self, section=None):
        # Allow both RequireSection(SECTION_X) instances and bare class use.
        if section is not None:
            self.section = section

    def __call__(self):
        # DRF instantiates permission classes; when we pass an instance we make
        # it callable so `RequireSection(SECTION_X)` works in permission_classes.
        return self

    def has_permission(self, request, view):
        return has_section(request.user, self.section)


class RequireAnySection(BasePermission):
    """Allow users who have AT LEAST ONE of the given sections.

    Usage: permission_classes = [IsAuthenticated, RequireAnySection(SECTION_A, SECTION_B)]
    """
    sections = ()
    message = 'Your account does not have access to this section.'

    def __init__(self, *sections):
        if sections:
            self.sections = sections

    def __call__(self):
        return self

    def has_permission(self, request, view):
        allowed = set(get_allowed_sections(request.user))
        return any(s in allowed for s in self.sections)


# Convenience: guards the raw expense ledger + all write/finance endpoints.
def _require_expenses():
    return RequireSection(SECTION_EXPENSES)


# Backwards-compatible alias — existing decorators reference BlockPnlOnly.
class BlockPnlOnly(BasePermission):
    """Deny access to endpoints that require the Expenses section."""
    message = 'Your account does not have access to this section.'

    def has_permission(self, request, view):
        return has_section(request.user, SECTION_EXPENSES)


class ExpensePagination(PageNumberPagination):
    """Pagination that exposes `page_size` in every response — lets the
    frontend compute total pages without hardcoding the size."""

    def get_paginated_response(self, data):
        return Response({
            'count': self.page.paginator.count,
            'next': self.get_next_link(),
            'previous': self.get_previous_link(),
            'page_size': self.get_page_size(self.request),
            'results': data,
        })


class BranchViewSet(viewsets.ModelViewSet):
    """CRUD for branches."""
    queryset = Branch.objects.all()
    serializer_class = BranchSerializer
    pagination_class = None
    permission_classes = [IsAuthenticated, RequireAnySection(SECTION_EXPENSES, SECTION_PETTYCASH)]


# Sentinel value for the payment-mode filter that selects entries whose mode was
# left blank (not entered) on the side that carries the amount.
NO_PAYMENT_MODE = '__none__'


def _blank_mode_q():
    """Q matching expenses with no payment mode on their active (amount) side."""
    return (
        (Q(debited_amount__gt=0) & Q(debit_payment_mode='')) |
        (Q(credited_amount__gt=0) & Q(credit_payment_mode=''))
    )


def _apply_mode_filter(qs, mode):
    """Apply the payment-mode filter: NO_PAYMENT_MODE → blank-mode entries,
    otherwise a case-insensitive match on either side."""
    if mode == NO_PAYMENT_MODE:
        return qs.filter(_blank_mode_q())
    return qs.filter(Q(credit_payment_mode__iexact=mode) | Q(debit_payment_mode__iexact=mode))


class ExpenseViewSet(viewsets.ModelViewSet):
    """CRUD for expenses with filtering and running balance."""

    pagination_class = ExpensePagination
    permission_classes = [IsAuthenticated, BlockPnlOnly]

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return ExpenseCreateSerializer
        return ExpenseSerializer

    def get_queryset(self):
        qs = Expense.objects.select_related('branch').all()

        # Filter by branch (can be ID or location name)
        branch_val = self.request.query_params.get('branch')
        if branch_val:
            if branch_val.isdigit():
                qs = qs.filter(branch_id=branch_val)
            else:
                qs = qs.filter(branch__location__icontains=branch_val)

        # Filter by category
        category = self.request.query_params.get('category')
        if category:
            qs = qs.filter(category__iexact=category)

        # Filter by date range
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)

        # Filter by payment mode (matches entries whose credit OR debit mode is it;
        # the NO_PAYMENT_MODE sentinel selects blank-mode entries).
        mode = self.request.query_params.get('payment_mode')
        if mode:
            qs = _apply_mode_filter(qs, mode)

        # Search
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(credit_remark__icontains=search) |
                Q(debit_remark__icontains=search) |
                Q(credit_person__icontains=search) |
                Q(debit_person__icontains=search) |
                Q(category__icontains=search)
            )

        return qs.order_by('date', 'created_at')

    def list(self, request, *args, **kwargs):
        """Override list to include running balance computation."""
        queryset = self.get_queryset()

        # Calculate running balance
        expenses = list(queryset)
        initial_balances = {m.payment_mode: m.initial_balance for m in PaymentModeBalance.objects.all()}
        balances = {}
        for expense in expenses:
            credit = expense.credited_amount or Decimal('0.00')
            debit = expense.debited_amount or Decimal('0.00')
            
            if debit > 0:
                mode = expense.debit_payment_mode or 'Other'
                if mode not in balances:
                    balances[mode] = initial_balances.get(mode, Decimal('0.00'))
                balances[mode] -= debit
                
            if credit > 0:
                mode = expense.credit_payment_mode or 'Other'
                if mode not in balances:
                    balances[mode] = initial_balances.get(mode, Decimal('0.00'))
                balances[mode] += credit
                
            expense.running_balances = balances.copy()

        # Reconciliation vs the bank statements: flag each bank-mode expense as
        # present ('in_statement') or missing ('not_in_statement') from its bank's
        # statement. Non-bank modes (Cash, UPI, …) get None (no badge shown).
        stmt_by_bank = {
            b: _bank_statement_by_expense(b)
            for b in (BankStatementEntry.BANK_IDFC, BankStatementEntry.BANK_BOB)
        }
        stmt_summary = {'in_statement': 0, 'not_in_statement': 0}
        for expense in expenses:
            mode = (expense.debit_payment_mode if (expense.debited_amount or Decimal('0.00')) > 0
                    else expense.credit_payment_mode)
            bank = _bank_for_mode(mode)
            if bank:
                stmt = stmt_by_bank[bank].get(expense.id)
                expense.statement_status = 'in_statement' if stmt else 'not_in_statement'
                expense.matched_statement = stmt
                stmt_summary[expense.statement_status] += 1
            else:
                expense.statement_status = None
                expense.matched_statement = None

        # To show newest first, reverse the list AFTER calculating running balances,
        # then paginate. This keeps correct balance cache on each object.
        expenses.reverse()

        # Pagination
        page = self.paginate_queryset(expenses)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            resp = self.get_paginated_response(serializer.data)
            # Reconciliation totals across the FULL filtered set (not just this page),
            # so the Expenses section can show a "Matched with Statement" strip.
            resp.data['statement_summary'] = stmt_summary
            return resp

        serializer = self.get_serializer(expenses, many=True)
        return Response({'results': serializer.data, 'statement_summary': stmt_summary})

    @action(detail=False, methods=['delete'], url_path='delete-all')
    def delete_all(self, request):
        """Delete all expenses. Gated by the admin-configured clear-data password
        so it can't be triggered by accident or via a direct API call."""
        setting = AppSetting.get_solo()
        if not setting.clear_password_is_set:
            return Response(
                {'detail': 'No clear-data password is configured. Ask an admin to set one in Settings first.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        password = request.data.get('password') if isinstance(request.data, dict) else None
        if not password or not setting.check_clear_password(password):
            return Response({'detail': 'Incorrect password.'}, status=status.HTTP_403_FORBIDDEN)

        count, _ = Expense.objects.all().delete()
        return Response({'detail': f'Successfully deleted {count} expenses.'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='filter-options')
    def filter_options(self, request):
        """Branches and categories ACTUALLY used in expense entries — for the
        filter dropdowns. De-duplicated case-insensitively, junk removed, and
        title-cased for display. Backend filters via icontains/iexact so a single
        cleaned entry matches every casing variant."""
        def clean(values, junk=frozenset()):
            seen = {}
            for v in values:
                v = ' '.join((v or '').split())
                if not v:
                    continue
                key = v.upper()
                if key in junk:
                    continue
                if key not in seen:
                    seen[key] = ' '.join(w[:1].upper() + w[1:].lower() for w in v.split(' '))
            return sorted(seen.values(), key=lambda s: s.lower())

        branch_vals = (
            Expense.objects.exclude(branch__location='')
            .values_list('branch__location', flat=True).distinct()
        )
        cat_vals = (
            Expense.objects.exclude(category='')
            .values_list('category', flat=True).distinct()
        )
        return Response({
            'branches': clean(branch_vals, frozenset({'NULL', 'NONE', 'UNDEFINED'})),
            'categories': clean(cat_vals),
        })


@api_view(['GET'])
@permission_classes([IsAuthenticated, RequireAnySection(SECTION_DASHBOARD, SECTION_REGION)])
def dashboard_view(request):
    """Aggregated dashboard stats. Also feeds the Region Expense page, so users
    with either the Dashboard or Region section may read it."""
    qs = Expense.objects.all()

    # Apply same filters
    branch_val = request.query_params.get('branch')
    if branch_val:
        if branch_val.isdigit():
            qs = qs.filter(branch_id=branch_val)
        else:
            qs = qs.filter(branch__location__icontains=branch_val)

    category = request.query_params.get('category')
    if category:
        qs = qs.filter(category__iexact=category)

    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    mode = request.query_params.get('payment_mode')
    if mode:
        qs = _apply_mode_filter(qs, mode)

    totals = qs.aggregate(
        total_credits=Coalesce(Sum('credited_amount'), Decimal('0.00')),
        total_debits=Coalesce(Sum('debited_amount'), Decimal('0.00')),
    )

    if mode == NO_PAYMENT_MODE:
        # Blank-mode entries have no configured balance; report their net movement.
        blank_qs = Expense.objects.filter(_blank_mode_q())
        mode_credits = blank_qs.aggregate(t=Coalesce(Sum('credited_amount'), Decimal('0.00')))['t']
        mode_debits = blank_qs.aggregate(t=Coalesce(Sum('debited_amount'), Decimal('0.00')))['t']
        total_balance = mode_credits - mode_debits
    elif mode:
        # Balance of the selected payment mode (all-time, cumulative).
        mode_initial = PaymentModeBalance.objects.filter(payment_mode__iexact=mode).aggregate(t=Coalesce(Sum('initial_balance'), Decimal('0.00')))['t']
        mode_credits = Expense.objects.filter(credit_payment_mode__iexact=mode).aggregate(t=Coalesce(Sum('credited_amount'), Decimal('0.00')))['t']
        mode_debits = Expense.objects.filter(debit_payment_mode__iexact=mode).aggregate(t=Coalesce(Sum('debited_amount'), Decimal('0.00')))['t']
        total_balance = mode_initial + mode_credits - mode_debits
    else:
        # Total Balance is the sum of all Payment Mode balances (company-wide).
        total_initial = PaymentModeBalance.objects.aggregate(t=Coalesce(Sum('initial_balance'), Decimal('0.00')))['t']
        global_credits = Expense.objects.aggregate(t=Coalesce(Sum('credited_amount'), Decimal('0.00')))['t']
        global_debits = Expense.objects.aggregate(t=Coalesce(Sum('debited_amount'), Decimal('0.00')))['t']
        total_balance = total_initial + global_credits - global_debits

    # Category breakdown (for pie chart)
    category_data = (
        qs.values('category')
        .annotate(
            total_credit=Coalesce(Sum('credited_amount'), Decimal('0.00')),
            total_debit=Coalesce(Sum('debited_amount'), Decimal('0.00')),
        )
        .order_by('category')
    )

    # Monthly trend (for line chart)
    monthly_data = (
        qs.annotate(month=TruncMonth('date'))
        .values('month')
        .annotate(
            credits=Coalesce(Sum('credited_amount'), Decimal('0.00')),
            debits=Coalesce(Sum('debited_amount'), Decimal('0.00')),
        )
        .order_by('month')
    )

    # Branch-wise (for bar chart and detailed breakdown)
    branch_qs = (
        qs.values('branch__location')
        .annotate(
            total_credit=Coalesce(Sum('credited_amount'), Decimal('0.00')),
            total_debit=Coalesce(Sum('debited_amount'), Decimal('0.00')),
        )
        .order_by('branch__location')
    )

    branch_breakdown = []
    for item in branch_qs:
        location = item['branch__location']
        # Category breakdown for THIS branch (scoped to current filters)
        branch_cats = (
            qs.filter(branch__location=location)
            .values('category')
            .annotate(
                total_credit=Coalesce(Sum('credited_amount'), Decimal('0.00')),
                total_debit=Coalesce(Sum('debited_amount'), Decimal('0.00')),
            )
            .order_by('-total_debit')
        )

        branch_breakdown.append({
            'branch': location,
            'total_credit': str(item['total_credit']),
            'total_debit': str(item['total_debit']),
            'category_breakdown': [
                {
                    'category': c['category'],
                    'total_credit': str(c['total_credit']),
                    'total_debit': str(c['total_debit']),
                }
                for c in branch_cats
            ]
        })

    return Response({
        'total_balance': str(total_balance),
        'total_credits': str(totals['total_credits']),
        'total_debits': str(totals['total_debits']),
        'category_breakdown': [
            {
                'category': item['category'],
                'total_credit': str(item['total_credit']),
                'total_debit': str(item['total_debit']),
            }
            for item in category_data
        ],
        'monthly_trend': [
            {
                'month': item['month'].strftime('%Y-%m') if item['month'] else '',
                'credits': str(item['credits']),
                'debits': str(item['debits']),
            }
            for item in monthly_data
        ],
        'branch_breakdown': branch_breakdown,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, BlockPnlOnly])
def export_expenses(request):
    """Export expenses to CSV."""
    qs = Expense.objects.select_related('branch').all().order_by('date', 'created_at')

    # Apply filters
    branch_val = request.query_params.get('branch')
    if branch_val:
        if branch_val.isdigit():
            qs = qs.filter(branch_id=branch_val)
        else:
            qs = qs.filter(branch__location__icontains=branch_val)

    category = request.query_params.get('category')
    if category:
        qs = qs.filter(category__iexact=category)

    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    mode = request.query_params.get('payment_mode')
    if mode:
        qs = _apply_mode_filter(qs, mode)

    # Note: avoid the name `format` — DRF reserves it for content negotiation.
    fmt = request.query_params.get('type', 'csv')

    if fmt == 'excel':
        # Excel export
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Expenses'

            headers = ['S.No', 'Date', 'Category', 'Branch', 'Credit Amount',
                        'Credit Remark', 'Credit Person', 'Credit Payment Mode',
                        'Debit Amount', 'Debit Remark', 'Debit Person',
                        'Debit Payment Mode', 'Running Balance']
            ws.append(headers)

            initial_balances = {m.payment_mode: m.initial_balance for m in PaymentModeBalance.objects.all()}
            balances = {}
            for idx, expense in enumerate(qs, 1):
                credit = expense.credited_amount or Decimal('0.00')
                debit = expense.debited_amount or Decimal('0.00')
                if debit > 0:
                    mode = expense.debit_payment_mode or 'Other'
                    if mode not in balances:
                        balances[mode] = initial_balances.get(mode, Decimal('0.00'))
                    balances[mode] -= debit
                    
                if credit > 0:
                    mode = expense.credit_payment_mode or 'Other'
                    if mode not in balances:
                        balances[mode] = initial_balances.get(mode, Decimal('0.00'))
                    balances[mode] += credit

                running_balance = " | ".join(f"{k}: {float(v)}" for k, v in balances.items() if v != Decimal('0.00'))


                ws.append([
                    idx,
                    expense.date.strftime('%Y-%m-%d'),
                    expense.category,
                    expense.branch.location,
                    float(credit),
                    expense.credit_remark,
                    expense.credit_person,
                    expense.credit_payment_mode,
                    float(debit),
                    expense.debit_remark,
                    expense.debit_person,
                    expense.debit_payment_mode,
                    running_balance,
                ])

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="expenses.xlsx"'
            return response
        except ImportError:
            return Response(
                {"error": "openpyxl not installed for Excel export"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    else:
        # CSV export
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="expenses.csv"'

        writer = csv.writer(response)
        writer.writerow(['S.No', 'Date', 'Category', 'Branch', 'Credit Amount',
                          'Credit Remark', 'Credit Person', 'Credit Payment Mode',
                          'Debit Amount', 'Debit Remark', 'Debit Person',
                          'Debit Payment Mode', 'Running Balance'])

        initial_balances = {m.payment_mode: m.initial_balance for m in PaymentModeBalance.objects.all()}
        balances = {}
        for idx, expense in enumerate(qs, 1):
            credit = expense.credited_amount or Decimal('0.00')
            debit = expense.debited_amount or Decimal('0.00')
            if debit > 0:
                mode = expense.debit_payment_mode or 'Other'
                if mode not in balances:
                    balances[mode] = initial_balances.get(mode, Decimal('0.00'))
                balances[mode] -= debit
                
            if credit > 0:
                mode = expense.credit_payment_mode or 'Other'
                if mode not in balances:
                    balances[mode] = initial_balances.get(mode, Decimal('0.00'))
                balances[mode] += credit

            running_balance = " | ".join(f"{k}: {float(v)}" for k, v in balances.items() if v != Decimal('0.00'))


            writer.writerow([
                idx,
                expense.date.strftime('%Y-%m-%d'),
                expense.category,
                expense.branch.location,
                credit,
                expense.credit_remark,
                expense.credit_person,
                expense.credit_payment_mode,
                debit,
                expense.debit_remark,
                expense.debit_person,
                expense.debit_payment_mode,
                running_balance,
            ])

        return response


# ---------------------------------------------------------------------------
# Payment Mode Balances
# ---------------------------------------------------------------------------
@api_view(['GET'])
@permission_classes([IsAuthenticated, BlockPnlOnly])
def payment_mode_balances_view(request):
    """Return all payment modes with initial + current balance, including those only present in expenses.
    
    Supports filtering:
      - fy: Financial year, e.g. '2025-2026' (April to March)
      - date_from / date_to: Custom date range
    """
    # Build expense filter based on query params
    expense_filter = Q()
    fy = request.query_params.get('fy')
    month = request.query_params.get('month')
    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')

    start_date = None
    end_date = None

    if fy:
        try:
            parts = fy.split('-')
            start_year = int(parts[0])
            end_year = int(parts[1])
            
            if month:
                # month is 1-12
                m = int(month)
                # In India, FY starts in April. 
                # April (4) to Dec (12) are in start_year. 
                # Jan (1) to March (3) are in end_year.
                year = start_year if m >= 4 else end_year
                import calendar
                _, last_day = calendar.monthrange(year, m)
                start_date = f'{year}-{m:02d}-01'
                end_date = f'{year}-{m:02d}-{last_day:02d}'
            else:
                start_date = f'{start_year}-04-01'
                end_date = f'{end_year}-03-31'
            
            expense_filter &= Q(date__gte=start_date, date__lte=end_date)
        except (ValueError, IndexError):
            pass
    elif month:
        try:
            m = int(month)
            year = datetime.now().year
            import calendar
            _, last_day = calendar.monthrange(year, m)
            start_date = f'{year}-{m:02d}-01'
            end_date = f'{year}-{m:02d}-{last_day:02d}'
            expense_filter &= Q(date__gte=start_date, date__lte=end_date)
        except ValueError:
            pass
    else:
        if date_from:
            expense_filter &= Q(date__gte=date_from)
            start_date = date_from
        if date_to:
            expense_filter &= Q(date__lte=date_to)
            end_date = date_to

    balances = list(PaymentModeBalance.objects.all())
    explicit_modes = {b.payment_mode for b in balances}
    
    credit_modes = Expense.objects.exclude(credit_payment_mode='').values_list('credit_payment_mode', flat=True).distinct()
    debit_modes = Expense.objects.exclude(debit_payment_mode='').values_list('debit_payment_mode', flat=True).distinct()
    
    all_used_modes = set(credit_modes).union(set(debit_modes))
    missing_modes = all_used_modes - explicit_modes
    
    for mode in missing_modes:
        if mode:
            balances.append(PaymentModeBalance(
                id=len(balances) + 9999, # Fake ID to bypass frontend unique key warnings
                payment_mode=mode,
                initial_balance=Decimal('0.00')
            ))

    result = []
    for bal in balances:
        mode = bal.payment_mode
        
        # Calculate actual initial balance taking into account transactions before start_date
        period_initial = bal.initial_balance
        if start_date:
            past_credits = Expense.objects.filter(
                date__lt=start_date, credit_payment_mode=mode
            ).aggregate(
                total=Coalesce(Sum('credited_amount'), Decimal('0.00'))
            )['total']
            past_debits = Expense.objects.filter(
                date__lt=start_date, debit_payment_mode=mode
            ).aggregate(
                total=Coalesce(Sum('debited_amount'), Decimal('0.00'))
            )['total']
            period_initial += past_credits - past_debits

        # Credits with this payment mode (filtered)
        total_credits = Expense.objects.filter(
            expense_filter, credit_payment_mode=mode
        ).aggregate(
            total=Coalesce(Sum('credited_amount'), Decimal('0.00'))
        )['total']
        # Debits with this payment mode (filtered)
        total_debits = Expense.objects.filter(
            expense_filter, debit_payment_mode=mode
        ).aggregate(
            total=Coalesce(Sum('debited_amount'), Decimal('0.00'))
        )['total']
        
        current = period_initial + total_credits - total_debits
        period_available = total_credits - total_debits

        bal.initial_balance = period_initial
        bal.current_balance = current
        bal.total_credits = total_credits
        bal.total_debits = total_debits
        bal.period_available = period_available
        
        result.append(bal)

    serializer = PaymentModeBalanceSerializer(result, many=True)
    return Response(serializer.data)



@api_view(['POST'])
@permission_classes([IsAuthenticated, BlockPnlOnly])
def payment_mode_balance_set(request):
    """Create or update initial balance for a payment mode."""
    mode = request.data.get('payment_mode', '').strip()
    initial = request.data.get('initial_balance')

    if not mode:
        return Response(
            {'detail': 'payment_mode is required.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    obj, created = PaymentModeBalance.objects.update_or_create(
        payment_mode=mode,
        defaults={'initial_balance': Decimal(str(initial or 0))},
    )

    # Compute current balance
    total_credits = Expense.objects.filter(
        credit_payment_mode=mode
    ).aggregate(
        total=Coalesce(Sum('credited_amount'), Decimal('0.00'))
    )['total']
    total_debits = Expense.objects.filter(
        debit_payment_mode=mode
    ).aggregate(
        total=Coalesce(Sum('debited_amount'), Decimal('0.00'))
    )['total']
    obj.current_balance = obj.initial_balance + total_credits - total_debits

    serializer = PaymentModeBalanceSerializer(obj)
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated, BlockPnlOnly])
def payment_mode_balance_delete(request):
    """Delete a payment mode balance entry."""
    mode = request.data.get('payment_mode', '').strip()

    if not mode:
        return Response(
            {'detail': 'payment_mode is required.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        obj = PaymentModeBalance.objects.get(payment_mode=mode)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    except PaymentModeBalance.DoesNotExist:
        return Response(
            {'detail': f'Payment mode "{mode}" not found.'},
            status=status.HTTP_404_NOT_FOUND,
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated, BlockPnlOnly])
def payment_mode_balance_rename(request):
    """Rename a payment mode EVERYWHERE — the PaymentModeBalance row and every
    expense entry that used the old name (both credit and debit mode fields),
    case-insensitively. If a mode with the new name already exists, the two are
    merged (its balance row is kept). All-or-nothing in one transaction."""
    from django.db import transaction
    old = (request.data.get('old_name') or '').strip()
    new = (request.data.get('new_name') or '').strip()
    if not old or not new:
        return Response({'detail': 'old_name and new_name are required.'}, status=status.HTTP_400_BAD_REQUEST)
    if old.lower() == new.lower():
        return Response({'detail': 'The new name is the same as the old name.'}, status=status.HTTP_400_BAD_REQUEST)

    with transaction.atomic():
        # Rewrite every expense that referenced the old mode name.
        n_credit = Expense.objects.filter(credit_payment_mode__iexact=old).update(credit_payment_mode=new)
        n_debit = Expense.objects.filter(debit_payment_mode__iexact=old).update(debit_payment_mode=new)

        # Rename (or merge) the balance row.
        old_obj = PaymentModeBalance.objects.filter(payment_mode__iexact=old).first()
        new_obj = PaymentModeBalance.objects.filter(payment_mode__iexact=new).first()
        if old_obj:
            if new_obj and new_obj.pk != old_obj.pk:
                old_obj.delete()  # merge into the existing "new" row
            else:
                old_obj.payment_mode = new
                old_obj.save(update_fields=['payment_mode'])

    return Response({
        'detail': f'Renamed "{old}" → "{new}".',
        'updated_entries': n_credit + n_debit,
    }, status=status.HTTP_200_OK)


# ---------------------------------------------------------------------------
# Categories — expose model choices so the frontend doesn't hardcode them.
# ---------------------------------------------------------------------------
@api_view(['GET'])
def categories_view(request):
    """Expense categories: the model's suggested choices PLUS every distinct
    category actually used in the ledger, so filters/suggestions reflect real
    data. De-duplicated case-insensitively (first-seen casing wins)."""
    model_choices = [value for value, _ in Expense.CATEGORY_CHOICES]
    used = Expense.objects.exclude(category='').values_list('category', flat=True).distinct()
    seen = {}
    for c in list(model_choices) + list(used):
        c = ' '.join((c or '').split())
        if not c:
            continue
        key = c.upper()
        if key not in seen:
            seen[key] = c
    return Response(sorted(seen.values(), key=lambda s: s.lower()))


# ---------------------------------------------------------------------------
# Profit & Loss report — spreadsheet-style Income/Expense × Month matrix.
# ---------------------------------------------------------------------------

# Junk / aggregate "branch" values that should never appear as a real branch in
# the P&L. Compared case-insensitively after trimming.
_PNL_EXCLUDED_BRANCHES = {'', 'null', 'none', 'common', 'all location', 'all locations', 'main branch'}

# Expense category (canonical UPPER) -> parent group, so the P&L can show
# collapsible expense groups with per-group subtotals. Unmapped categories fall
# into 'Office & Others'.
_PNL_EXPENSE_GROUPS = {
    'SALARY': 'Salary & Wages', 'SALARY MONTHLY': 'Salary & Wages', 'SALARY ADVANCE': 'Salary & Wages',
    'SALARY SETTLEMENT': 'Salary & Wages', 'FF SETTLEMENT': 'Salary & Wages',
    'OFFICE RENT': 'Rent & Facilities', 'RENT': 'Rent & Facilities', 'ROOM RENT': 'Rent & Facilities', 'HOUSE KEEPING': 'Rent & Facilities',
    'EB BILL': 'Utilities & Recharge', 'WIFI RECHARGE': 'Utilities & Recharge', 'PHONE RECHARGE': 'Utilities & Recharge',
    'PETROL ALLOWANCE': 'Travel & Fuel', 'PETROL ADVANCE': 'Travel & Fuel', 'TRAVEL': 'Travel & Fuel',
    'TRAVEL EXPENSE': 'Travel & Fuel', 'TRAVEL EXPENSES': 'Travel & Fuel', 'CAB': 'Travel & Fuel', 'RAPIDO': 'Travel & Fuel', 'TRANSPORT': 'Travel & Fuel',
    'GST': 'Statutory & Fees', 'GST FEE': 'Statutory & Fees', 'AUDITOR FEES': 'Statutory & Fees',
    'LOAN RETURN': 'Loans & EMI', 'LAON': 'Loans & EMI', 'EMI': 'Loans & EMI', 'LAPTOP EMI': 'Loans & EMI',
    'ASSET': 'Purchases & Assets', 'ASST PURCHASE': 'Purchases & Assets', 'PURCHASE': 'Purchases & Assets',
    'TOOL PURCHASE': 'Purchases & Assets', 'SOFTWARE PURCHASE': 'Purchases & Assets', 'SERVER PURCHASE': 'Purchases & Assets',
    'LYSTLOC': 'Subscriptions & Licenses', 'GPS LICENSE': 'Subscriptions & Licenses', 'FB LEAD': 'Subscriptions & Licenses',
    'COURIER': 'Office & Others', 'PETTY CASH': 'Office & Others', 'STATIONERY EXPENSES': 'Office & Others', 'STATIONARY': 'Office & Others',
    'VENDOR PAYMENT': 'Office & Others', 'REFUND': 'Office & Others', 'OTHER EXPENSE': 'Office & Others', 'OTHER EXPENSES': 'Office & Others', 'EXPENSES': 'Office & Others',
}
_PNL_DEFAULT_EXPENSE_GROUP = 'Office & Others'
# Display order of the groups (unlisted groups sort after these, alphabetically).
_PNL_GROUP_ORDER = [
    'Salary & Wages', 'Rent & Facilities', 'Utilities & Recharge', 'Travel & Fuel',
    'Statutory & Fees', 'Loans & EMI', 'Purchases & Assets', 'Subscriptions & Licenses', 'Office & Others',
]

# Categories that are always income even if a stray debit exists on them. These
# are the credit-driven revenue streams seen in the ledger. Matched on the
# canonical (UPPER + trimmed) category name.
_PNL_INCOME_CATEGORIES = {
    'LOAN', 'TRADE', 'SERVICE', 'EB RETURN', 'CLIENT RETURN', 'QUALITY TAX',
    'WARRANTY BILLS', 'WARRANTY BILL', 'BENCH ARC', 'ULR', 'OUT OF WARRANTY',
}

# Categories that are always expense even if a stray credit exists on them
# (e.g. a refund credited back onto SALARY/EXPENSES). Matched on canonical name.
_PNL_EXPENSE_CATEGORIES = {
    'SALARY', 'EXPENSES', 'EXPENSE', 'VENDOR', 'MISC', 'PETROL', 'PETROL ALLOWANCE',
    'RENT', 'OFFICE RENT', 'GST', 'LOAN RETURN',
}


def _pnl_canonical(name):
    """Canonical display form for a messy free-text label: trim + collapse
    internal whitespace + upper-case for matching. Returns '' for blank."""
    if not name:
        return ''
    return ' '.join(str(name).split()).upper()


def _pnl_classify(canonical_category, total_credit, total_debit):
    """Decide whether a category is 'income' or 'expense'.

    Data-driven with a hardcoded override for known ambiguous categories, so new
    categories added later are still classified automatically:
      1. Explicit income/expense sets win first.
      2. Otherwise, whichever side (credit vs debit) has the larger magnitude.
      3. Ties / all-zero fall back to 'expense' (the common case).
    """
    if canonical_category in _PNL_INCOME_CATEGORIES:
        return 'income'
    if canonical_category in _PNL_EXPENSE_CATEGORIES:
        return 'expense'
    if total_credit > total_debit:
        return 'income'
    return 'expense'


def _pnl_financial_year_bounds(fy_start_year):
    """Given a starting year (e.g. 2025) return (start_date, end_date) for the
    Indian financial year Apr 1 <start> – Mar 31 <start+1>, plus the ordered list
    of 12 month keys 'YYYY-MM' from April to March."""
    from datetime import date
    start = date(fy_start_year, 4, 1)
    end = date(fy_start_year + 1, 3, 31)
    months = []
    for i in range(12):
        m = 4 + i
        y = fy_start_year + (m - 1) // 12
        mm = (m - 1) % 12 + 1
        months.append(f'{y:04d}-{mm:02d}')
    return start, end, months


@api_view(['GET'])
@permission_classes([IsAuthenticated, RequireSection(SECTION_PNL)])
def profit_loss_view(request):
    """Profit & Loss matrix: income & expense categories (rows) × months (cols).

    Query params:
      fy       – financial-year start year, e.g. '2025' for FY 2025-26 (Apr–Mar).
                 Defaults to the FY containing the most recent expense (or today).
      branch   – branch id or (case-insensitive) location substring; junk/aggregate
                 branches are excluded from the default (all-branch) view.

    All messy real-world data is normalised in Python: branches and categories are
    merged case-insensitively, and income/expense is decided per category.
    """
    qs = Expense.objects.all()

    # ---- Branch filter (case-insensitive; digits => exact id) ----
    branch_val = request.query_params.get('branch')
    if branch_val:
        if branch_val.isdigit():
            qs = qs.filter(branch_id=branch_val)
        else:
            qs = qs.filter(branch__location__icontains=branch_val)

    # ---- Payment-mode filter (case-insensitive; matches either credit or debit side) ----
    # Lets the P&L be scoped to a single mode, e.g. only "Bank of Baroda" entries.
    mode_val = request.query_params.get('payment_mode')
    if mode_val:
        qs = qs.filter(Q(credit_payment_mode__iexact=mode_val) | Q(debit_payment_mode__iexact=mode_val))

    # ---- Determine the financial year window ----
    fy_param = request.query_params.get('fy')
    if fy_param and fy_param.isdigit():
        fy_start = int(fy_param)
    else:
        latest = Expense.objects.order_by('-date').values_list('date', flat=True).first()
        ref = latest if latest else datetime.now().date()
        # Jan–Mar belongs to the FY that started the previous calendar year.
        fy_start = ref.year if ref.month >= 4 else ref.year - 1

    start_date, end_date, month_keys = _pnl_financial_year_bounds(fy_start)
    qs = qs.filter(date__gte=start_date, date__lte=end_date)

    # ---- Aggregate: category × month, credit & debit ----
    rows = (
        qs.annotate(month=TruncMonth('date'))
        .values('category', 'month')
        .annotate(
            credit=Coalesce(Sum('credited_amount'), Decimal('0.00')),
            debit=Coalesce(Sum('debited_amount'), Decimal('0.00')),
        )
    )

    # Merge case-variant categories together and bucket amounts by month.
    # canon -> {'display': str, 'credit_total': Decimal, 'debit_total': Decimal,
    #           'months': {month_key: {'credit': Decimal, 'debit': Decimal}}}
    cat_map = {}
    for r in rows:
        canon = _pnl_canonical(r['category'])
        if not canon:
            canon = 'UNCATEGORISED'
        month_key = r['month'].strftime('%Y-%m') if r['month'] else ''
        entry = cat_map.get(canon)
        if entry is None:
            # Title-case a nicer display label from the canonical form.
            entry = {
                'display': canon.title(),
                'credit_total': Decimal('0.00'),
                'debit_total': Decimal('0.00'),
                'months': {},
            }
            cat_map[canon] = entry
        entry['credit_total'] += r['credit']
        entry['debit_total'] += r['debit']
        mslot = entry['months'].setdefault(month_key, {'credit': Decimal('0.00'), 'debit': Decimal('0.00')})
        mslot['credit'] += r['credit']
        mslot['debit'] += r['debit']

    # Build income & expense row lists. For income rows the per-month value is the
    # credit; for expense rows it's the debit — that's what a P&L shows.
    income_rows, expense_rows = [], []
    for canon, entry in cat_map.items():
        kind = _pnl_classify(canon, entry['credit_total'], entry['debit_total'])
        if kind == 'income':
            monthly = {mk: entry['months'].get(mk, {}).get('credit', Decimal('0.00')) for mk in month_keys}
            total = sum(monthly.values(), Decimal('0.00'))
            if total == 0:
                continue  # skip empty income rows for this FY
            income_rows.append({'category': entry['display'], 'monthly': monthly, 'total': total})
        else:
            monthly = {mk: entry['months'].get(mk, {}).get('debit', Decimal('0.00')) for mk in month_keys}
            total = sum(monthly.values(), Decimal('0.00'))
            if total == 0:
                continue
            expense_rows.append({
                'category': entry['display'], 'monthly': monthly, 'total': total,
                'group': _PNL_EXPENSE_GROUPS.get(canon, _PNL_DEFAULT_EXPENSE_GROUP),
            })

    # Sort each section by grand total, largest first (most material rows on top).
    income_rows.sort(key=lambda x: x['total'], reverse=True)
    expense_rows.sort(key=lambda x: x['total'], reverse=True)

    def _serialize_rows(section):
        return [
            {
                'category': row['category'],
                'monthly': {mk: str(row['monthly'][mk]) for mk in month_keys},
                'total': str(row['total']),
                **({'group': row['group']} if 'group' in row else {}),
            }
            for row in section
        ]

    # Column (per-month) totals + net profit/loss per month.
    income_by_month = {mk: sum((row['monthly'][mk] for row in income_rows), Decimal('0.00')) for mk in month_keys}
    expense_by_month = {mk: sum((row['monthly'][mk] for row in expense_rows), Decimal('0.00')) for mk in month_keys}
    net_by_month = {mk: income_by_month[mk] - expense_by_month[mk] for mk in month_keys}

    total_income = sum(income_by_month.values(), Decimal('0.00'))
    total_expense = sum(expense_by_month.values(), Decimal('0.00'))

    # Build the branch dropdown list: canonical, de-duplicated, junk excluded.
    branch_names = {}
    for loc in Branch.objects.values_list('location', flat=True):
        canon = _pnl_canonical(loc)
        if canon.lower() in _PNL_EXCLUDED_BRANCHES:
            continue
        branch_names.setdefault(canon, canon.title())
    branches = sorted(branch_names.values())

    # Available financial years (for the FY picker), derived from data.
    first_date = Expense.objects.order_by('date').values_list('date', flat=True).first()
    last_date = Expense.objects.order_by('-date').values_list('date', flat=True).first()
    fy_list = []
    if first_date and last_date:
        lo_fy = first_date.year if first_date.month >= 4 else first_date.year - 1
        hi_fy = last_date.year if last_date.month >= 4 else last_date.year - 1
        fy_list = list(range(lo_fy, hi_fy + 1))
    if fy_start not in fy_list:
        fy_list.append(fy_start)
    fy_list = sorted(set(fy_list), reverse=True)

    return Response({
        'fy_start': fy_start,
        'fy_label': f'FY {fy_start}-{str(fy_start + 1)[-2:]}',
        'months': month_keys,
        'available_fys': fy_list,
        'branches': branches,
        'income': _serialize_rows(income_rows),
        'expense': _serialize_rows(expense_rows),
        'expense_group_order': _PNL_GROUP_ORDER,
        'income_by_month': {mk: str(income_by_month[mk]) for mk in month_keys},
        'expense_by_month': {mk: str(expense_by_month[mk]) for mk in month_keys},
        'net_by_month': {mk: str(net_by_month[mk]) for mk in month_keys},
        'total_income': str(total_income),
        'total_expense': str(total_expense),
        'net_profit': str(total_income - total_expense),
    })


# ---------------------------------------------------------------------------
# Business Insights — deterministic (no AI/LLM) analytics over the ledger:
# branch ranking, expense trends, a simple forecast, spike detection and
# rule-based recommendations. Reuses the P&L canonicalisation/classification.
# ---------------------------------------------------------------------------

def _insights_month_window(n):
    """The last `n` month keys ('YYYY-MM') ending at the most recent expense
    month (or the current month if there is no data), oldest first."""
    latest = Expense.objects.order_by('-date').values_list('date', flat=True).first()
    ref = latest if latest else datetime.now().date()
    y, m = ref.year, ref.month
    keys = []
    for _ in range(n):
        keys.append(f'{y:04d}-{m:02d}')
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    keys.reverse()
    return keys


def _next_month_key(mk):
    y, m = int(mk[:4]), int(mk[5:])
    m += 1
    if m == 13:
        m, y = 1, y + 1
    return f'{y:04d}-{m:02d}'


def _month_label(mk):
    names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    y, m = mk[:4], int(mk[5:])
    return f'{names[m - 1]} {y[2:]}' if 1 <= m <= 12 else mk


@api_view(['GET'])
@permission_classes([IsAuthenticated, RequireSection(SECTION_INSIGHTS)])
def insights_view(request):
    """AI-free business analytics.

    Query params (all optional):
      months     – size of the trailing window in months (3-24, default 12).
      date_from  – 'YYYY-MM-DD'; with date_to, overrides the trailing window.
      date_to    – 'YYYY-MM-DD'.
      branch     – branch id, or a case-insensitive location substring.
    """
    from calendar import monthrange

    def _parse_date(value):
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except (TypeError, ValueError):
            return None

    date_from = _parse_date(request.query_params.get('date_from'))
    date_to = _parse_date(request.query_params.get('date_to'))

    if date_from and date_to and date_from <= date_to:
        # Explicit range: bucket by month across the whole span.
        start, end = date_from, date_to
        month_keys, y, m = [], start.year, start.month
        while (y, m) <= (end.year, end.month):
            month_keys.append(f'{y:04d}-{m:02d}')
            m += 1
            if m == 13:
                m, y = 1, y + 1
        n = len(month_keys)
    else:
        try:
            n = max(3, min(24, int(request.query_params.get('months', 12))))
        except (TypeError, ValueError):
            n = 12
        month_keys = _insights_month_window(n)
        start = datetime.strptime(month_keys[0] + '-01', '%Y-%m-%d').date()
        ey, em = int(month_keys[-1][:4]), int(month_keys[-1][5:])
        end = datetime(ey, em, monthrange(ey, em)[1]).date()

    qs = Expense.objects.filter(date__gte=start, date__lte=end)

    # ---- Branch filter (case-insensitive; digits => exact id) ----
    branch_val = request.query_params.get('branch')
    if branch_val:
        if branch_val.isdigit():
            qs = qs.filter(branch_id=branch_val)
        else:
            qs = qs.filter(branch__location__icontains=branch_val)

    ZERO = Decimal('0.00')

    # ---- category × month aggregation → classify + monthly trend ----
    cm = (
        qs.annotate(month=TruncMonth('date'))
        .values('category', 'month')
        .annotate(credit=Coalesce(Sum('credited_amount'), ZERO), debit=Coalesce(Sum('debited_amount'), ZERO))
    )
    cats = {}  # canon -> {display, credit_total, debit_total, months:{mk:{credit,debit}}, group}
    for r in cm:
        canon = _pnl_canonical(r['category']) or 'UNCATEGORISED'
        e = cats.get(canon)
        if e is None:
            e = {'display': canon.title(), 'credit_total': ZERO, 'debit_total': ZERO, 'months': {}}
            cats[canon] = e
        e['credit_total'] += r['credit']
        e['debit_total'] += r['debit']
        mk = r['month'].strftime('%Y-%m') if r['month'] else ''
        slot = e['months'].setdefault(mk, {'credit': ZERO, 'debit': ZERO})
        slot['credit'] += r['credit']
        slot['debit'] += r['debit']

    cat_kind = {c: _pnl_classify(c, e['credit_total'], e['debit_total']) for c, e in cats.items()}

    inc_by_month = {mk: ZERO for mk in month_keys}
    exp_by_month = {mk: ZERO for mk in month_keys}
    expense_cats = []  # {canon, category, total, monthly, group}
    for canon, e in cats.items():
        if cat_kind[canon] == 'income':
            for mk in month_keys:
                inc_by_month[mk] += e['months'].get(mk, {}).get('credit', ZERO)
        else:
            monthly = {mk: e['months'].get(mk, {}).get('debit', ZERO) for mk in month_keys}
            for mk in month_keys:
                exp_by_month[mk] += monthly[mk]
            total = sum(monthly.values(), ZERO)
            if total > 0:
                expense_cats.append({
                    'canon': canon, 'category': e['display'], 'total': total, 'monthly': monthly,
                    'group': _PNL_EXPENSE_GROUPS.get(canon, _PNL_DEFAULT_EXPENSE_GROUP),
                })
    net_by_month = {mk: inc_by_month[mk] - exp_by_month[mk] for mk in month_keys}

    total_income = sum(inc_by_month.values(), ZERO)
    total_expense = sum(exp_by_month.values(), ZERO)
    net_profit = total_income - total_expense

    # ---- month-by-month table: each month's result, its biggest expense head,
    # and how the profit moved vs the month before ----
    monthly_breakdown = []
    prev_net = None
    for mk in month_keys:
        inc, exp = inc_by_month[mk], exp_by_month[mk]
        net = net_by_month[mk]
        top_cat, top_amt = None, ZERO
        for c in expense_cats:
            amt = c['monthly'].get(mk, ZERO)
            if amt > top_amt:
                top_cat, top_amt = c['category'], amt
        # Margin: what fraction of income was kept as profit.
        margin = round(float(net / inc * 100), 1) if inc > 0 else None
        change = None
        if prev_net is not None:
            change = str((net - prev_net).quantize(Decimal('0.01')))
        monthly_breakdown.append({
            'month': mk,
            'income': str(inc),
            'expense': str(exp),
            'net': str(net),
            'is_profit': net >= 0,
            'margin_pct': margin,
            'change_vs_prev': change,
            'top_expense': top_cat,
            'top_expense_amount': str(top_amt) if top_cat else None,
            'has_data': (inc > 0 or exp > 0),
        })
        prev_net = net

    # Best and worst months (only among months that actually have entries).
    active_months = [m for m in monthly_breakdown if m['has_data']]
    best_month = max(active_months, key=lambda m: Decimal(m['net']), default=None)
    worst_month = min(active_months, key=lambda m: Decimal(m['net']), default=None)
    profit_months = sum(1 for m in active_months if m['is_profit'])
    loss_months = len(active_months) - profit_months

    # ---- branch ranking (income vs expense per canonical branch) ----
    br = (
        qs.values('branch__location', 'category')
        .annotate(credit=Coalesce(Sum('credited_amount'), ZERO), debit=Coalesce(Sum('debited_amount'), ZERO))
    )
    branches = {}  # canon -> {display, income, expense}
    for r in br:
        cb = _pnl_canonical(r['branch__location'])
        if not cb or cb.lower() in _PNL_EXCLUDED_BRANCHES:
            continue
        b = branches.get(cb)
        if b is None:
            b = {'display': cb.title(), 'income': ZERO, 'expense': ZERO}
            branches[cb] = b
        canon_cat = _pnl_canonical(r['category']) or 'UNCATEGORISED'
        kind = cat_kind.get(canon_cat) or _pnl_classify(canon_cat, r['credit'], r['debit'])
        if kind == 'income':
            b['income'] += r['credit']
        else:
            b['expense'] += r['debit']
    branch_ranking = sorted(
        ({'branch': b['display'], 'income': b['income'], 'expense': b['expense'], 'net': b['income'] - b['expense']}
         for b in branches.values()),
        key=lambda x: x['net'], reverse=True,
    )

    # ---- forecast next month: average of the last 3 months ----
    tail = month_keys[-3:]
    fc_income = sum((inc_by_month[mk] for mk in tail), ZERO) / len(tail)
    fc_expense = sum((exp_by_month[mk] for mk in tail), ZERO) / len(tail)
    forecast = {
        'month': _next_month_key(month_keys[-1]),
        'income': fc_income.quantize(Decimal('0.01')),
        'expense': fc_expense.quantize(Decimal('0.01')),
        'net': (fc_income - fc_expense).quantize(Decimal('0.01')),
    }

    # ---- per-category recent growth (last 3 vs previous 3 months) ----
    def growth_pct(monthly):
        if len(month_keys) < 6:
            return None
        last3 = sum((monthly[mk] for mk in month_keys[-3:]), ZERO)
        prev3 = sum((monthly[mk] for mk in month_keys[-6:-3]), ZERO)
        if prev3 <= 0:
            return None
        return round(float((last3 - prev3) / prev3 * 100), 1)

    expense_cats.sort(key=lambda c: c['total'], reverse=True)
    top_expenses = [{
        'category': c['category'],
        'total': str(c['total']),
        'share': round(float(c['total'] / total_expense), 4) if total_expense > 0 else 0,
        'growth_pct': growth_pct(c['monthly']),
        'group': c['group'],
    } for c in expense_cats[:8]]

    # ---- anomalies: latest month >> its own prior-months average ----
    anomalies = []
    latest_mk = month_keys[-1]
    for c in expense_cats:
        prior = [c['monthly'][mk] for mk in month_keys[:-1]]
        nonzero = [v for v in prior if v > 0]
        if not nonzero:
            continue
        avg = sum(nonzero, ZERO) / len(nonzero)
        cur = c['monthly'][latest_mk]
        if avg > 0 and cur >= avg * Decimal('2') and cur >= Decimal('5000'):
            anomalies.append({
                'category': c['category'], 'month': latest_mk,
                'amount': str(cur), 'avg': str(avg.quantize(Decimal('0.01'))),
                'times': round(float(cur / avg), 1),
            })
    anomalies.sort(key=lambda a: float(a['amount']), reverse=True)

    # ---- rule-based recommendations / action points ----
    recs = []

    def money(v):
        return f'₹{float(v):,.0f}'

    # Headline: profit/loss stated with the margin, i.e. what is kept per ₹100 earned.
    if total_income > 0 or total_expense > 0:
        if net_profit >= 0:
            kept = f' You keep about {money(net_profit / total_income * 100)} out of every ₹100 you earn.' if total_income > 0 else ''
            recs.append({'kind': 'good', 'title': 'You are in profit',
                         'text': f'Across {len(active_months) or len(month_keys)} months you earned {money(total_income)} '
                                 f'and spent {money(total_expense)}, leaving a profit of {money(net_profit)}.{kept}'})
        else:
            overspend = f' For every ₹100 earned you are spending about {money(total_expense / total_income * 100)}.' if total_income > 0 else ''
            recs.append({'kind': 'alert', 'title': 'You are running at a loss',
                         'text': f'Across {len(active_months) or len(month_keys)} months you earned {money(total_income)} '
                                 f'but spent {money(total_expense)} — short by {money(-net_profit)}.{overspend} '
                                 f'The biggest expense heads below are where to look first.'})

    # How consistent is the result month to month?
    if len(active_months) >= 3:
        if loss_months == 0:
            recs.append({'kind': 'good', 'title': 'Every month was profitable',
                         'text': f'All {len(active_months)} months with entries ended in profit. '
                                 f'That is a steady business — keep the current cost pattern.'})
        elif profit_months == 0:
            recs.append({'kind': 'alert', 'title': 'No month was profitable',
                         'text': f'All {len(active_months)} months ended in loss. This is not a one-off month — '
                                 f'the cost structure itself needs review, not just one expense.'})
        else:
            recs.append({'kind': 'tip', 'title': f'{profit_months} profit months, {loss_months} loss months',
                         'text': f'Out of {len(active_months)} months, {profit_months} made money and {loss_months} lost money. '
                                 f'Compare a profit month against a loss month in the table below to see what changed.'})

    # Point at the best and worst months by name so they can be examined directly.
    if best_month and worst_month and best_month['month'] != worst_month['month']:
        recs.append({'kind': 'good', 'title': f'Best month: {_month_label(best_month["month"])}',
                     'text': f'{_month_label(best_month["month"])} was your strongest month at '
                             f'{money(float(best_month["net"]))} profit (income {money(float(best_month["income"]))}). '
                             f'Look at what was different that month.'})
        if Decimal(worst_month['net']) < 0:
            recs.append({'kind': 'alert', 'title': f'Worst month: {_month_label(worst_month["month"])}',
                         'text': f'{_month_label(worst_month["month"])} lost {money(-float(worst_month["net"]))}'
                                 + (f', mostly on {worst_month["top_expense"]} '
                                    f'({money(float(worst_month["top_expense_amount"]))}).' if worst_month['top_expense'] else '.')})

    if net_by_month[latest_mk] < 0:
        recs.append({'kind': 'alert', 'title': f'{_month_label(latest_mk)}: spent more than earned',
                     'text': f'This month expense was {money(exp_by_month[latest_mk])} vs income '
                             f'{money(inc_by_month[latest_mk])} — a {money(-net_by_month[latest_mk])} gap.'})

    if forecast['net'] < 0:
        recs.append({'kind': 'alert', 'title': f'Next month may run negative',
                     'text': f'Based on your last 3 months, {_month_label(forecast["month"])} is trending to a '
                             f'{money(-forecast["net"])} shortfall. Plan cash accordingly.'})

    if branch_ranking:
        best = branch_ranking[0]
        if best['net'] > 0:
            recs.append({'kind': 'good', 'title': f'{best["branch"]} is your best branch',
                         'text': f'{best["branch"]} made the most profit ({money(best["net"])}). '
                                 f'Consider what is working there and replicate it.'})
        # Only flag materially loss-making branches (ignore tiny data-quirk rows).
        loss_branches = [b for b in branch_ranking if b['net'] < Decimal('-10000')]
        for b in loss_branches[:2]:
            recs.append({'kind': 'alert', 'title': f'{b["branch"]} is loss-making',
                         'text': f'{b["branch"]} is down {money(-b["net"])} (income {money(b["income"])}, '
                                 f'expense {money(b["expense"])}). Review its costs.'})

    if top_expenses and total_expense > 0:
        top = top_expenses[0]
        if top['share'] >= 0.30:
            recs.append({'kind': 'tip', 'title': f'{top["category"]} dominates your spend',
                         'text': f'{top["category"]} alone is {round(top["share"] * 100)}% of total expense '
                                 f'({money(float(top["total"]))}). Small savings here move the needle most.'})
        for c in top_expenses:
            if c['growth_pct'] is not None and c['growth_pct'] >= 40 and float(c['total']) >= 10000:
                recs.append({'kind': 'alert', 'title': f'{c["category"]} rising fast',
                             'text': f'{c["category"]} is up {c["growth_pct"]}% vs the previous 3 months. '
                                     f'Check whether this is expected.'})
                break

    for a in anomalies[:2]:
        recs.append({'kind': 'alert', 'title': f'{a["category"]} spiked this month',
                     'text': f'{a["category"]} was {money(float(a["amount"]))} in {_month_label(a["month"])} — '
                             f'{a["times"]}× its usual {money(float(a["avg"]))}. Worth a look.'})

    # Branch dropdown list — canonical, de-duplicated, junk excluded (as in P&L).
    branch_names = {}
    for loc in Branch.objects.values_list('location', flat=True):
        canon = _pnl_canonical(loc)
        if canon.lower() in _PNL_EXCLUDED_BRANCHES:
            continue
        branch_names.setdefault(canon, canon.title())
    all_branches = sorted(branch_names.values())

    return Response({
        'window_months': month_keys,
        'window_label': f'{_month_label(month_keys[0])} – {_month_label(month_keys[-1])}',
        'date_from': start.strftime('%Y-%m-%d'),
        'date_to': end.strftime('%Y-%m-%d'),
        'branches': all_branches,
        'summary': {
            'total_income': str(total_income),
            'total_expense': str(total_expense),
            'net_profit': str(net_profit),
            'is_profit': net_profit >= 0,
            'margin_pct': round(float(net_profit / total_income * 100), 1) if total_income > 0 else None,
            'active_months': len(active_months),
            'profit_months': profit_months,
            'loss_months': loss_months,
            'best_month': best_month['month'] if best_month else None,
            'worst_month': worst_month['month'] if worst_month else None,
            'latest_month': latest_mk,
            'latest_income': str(inc_by_month[latest_mk]),
            'latest_expense': str(exp_by_month[latest_mk]),
            'latest_net': str(net_by_month[latest_mk]),
        },
        'monthly_breakdown': monthly_breakdown,
        'monthly_trend': [
            {'month': mk, 'income': str(inc_by_month[mk]), 'expense': str(exp_by_month[mk]), 'net': str(net_by_month[mk])}
            for mk in month_keys
        ],
        'forecast': {k: (str(v) if isinstance(v, Decimal) else v) for k, v in forecast.items()},
        'branch_ranking': [
            {'branch': b['branch'], 'income': str(b['income']), 'expense': str(b['expense']), 'net': str(b['net'])}
            for b in branch_ranking
        ],
        'top_expenses': top_expenses,
        'anomalies': anomalies,
        'recommendations': recs,
    })


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------
@api_view(['POST'])
@permission_classes([AllowAny])
def login_view(request):
    """Username + password → token. Used by the SPA login form."""
    username = (request.data.get('username') or '').strip()
    password = request.data.get('password') or ''

    if not username or not password:
        return Response(
            {'detail': 'Username and password are required.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    user = authenticate(request, username=username, password=password)
    if user is None or not user.is_active:
        return Response(
            {'detail': 'Invalid credentials.'},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    token, _ = Token.objects.get_or_create(user=user)
    return Response({
        'token': token.key,
        'username': user.username,
        'is_staff': user.is_staff,
        'is_admin': is_admin_user(user),
        'allowed_sections': get_allowed_sections(user),
        'pnl_only': is_pnl_only(user),
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_view(request):
    """Invalidate the caller's token."""
    Token.objects.filter(user=request.user).delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def me_view(request):
    """Return current user info — used to verify a stored token is still valid."""
    return Response({
        'username': request.user.username,
        'is_staff': request.user.is_staff,
        'is_admin': is_admin_user(request.user),
        'allowed_sections': get_allowed_sections(request.user),
        'pnl_only': is_pnl_only(request.user),
    })


# ---------------------------------------------------------------------------
# Admin: user management (create login accounts + control section access)
# ---------------------------------------------------------------------------
from django.contrib.auth.models import User as AuthUserModel


class IsAdminSection(BasePermission):
    """Only staff/superusers may manage users."""
    message = 'Only administrators can manage users.'

    def has_permission(self, request, view):
        return is_admin_user(request.user)


def _serialize_managed_user(user):
    """Public shape of a user row for the admin panel."""
    return {
        'id': user.id,
        'username': user.username,
        'is_admin': is_admin_user(user),
        'is_active': user.is_active,
        'allowed_sections': get_allowed_sections(user),
        'date_joined': user.date_joined.isoformat() if user.date_joined else None,
    }


def _clean_section_input(raw):
    """Validate a requested section list → canonical, de-duplicated subset."""
    if not isinstance(raw, list):
        return None
    wanted = {str(s).strip().lower() for s in raw}
    return [s for s in ALL_SECTIONS if s in wanted]


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminSection])
def admin_sections_view(request):
    """List the sections that can be granted, with labels (for the UI)."""
    from .models import SECTION_LABELS
    return Response([
        {'key': key, 'label': SECTION_LABELS.get(key, key)}
        for key in ALL_SECTIONS
    ])


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, IsAdminSection])
def clear_data_password_view(request):
    """Admin-only. GET reports whether a clear-data password is set; POST sets or
    changes it (hashed). This password gates the destructive Clear All Data action."""
    setting = AppSetting.get_solo()
    if request.method == 'GET':
        return Response({'is_set': setting.clear_password_is_set})

    password = (request.data.get('password') or '') if isinstance(request.data, dict) else ''
    if len(password) < 4:
        return Response({'detail': 'Password must be at least 4 characters.'}, status=status.HTTP_400_BAD_REQUEST)
    setting.set_clear_password(password)
    setting.save()
    return Response({'is_set': True})


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, IsAdminSection])
def admin_users_view(request):
    """GET: list all users. POST: create a new login account with sections."""
    if request.method == 'GET':
        users = AuthUserModel.objects.all().order_by('-is_superuser', 'username')
        return Response([_serialize_managed_user(u) for u in users])

    # --- POST: create ---
    username = (request.data.get('username') or '').strip()
    password = request.data.get('password') or ''
    sections = _clean_section_input(request.data.get('allowed_sections', []))

    if not username or not password:
        return Response({'detail': 'Username and password are required.'}, status=status.HTTP_400_BAD_REQUEST)
    if sections is None:
        return Response({'detail': 'allowed_sections must be a list.'}, status=status.HTTP_400_BAD_REQUEST)
    if AuthUserModel.objects.filter(username__iexact=username).exists():
        return Response({'detail': f"A user named '{username}' already exists."}, status=status.HTTP_400_BAD_REQUEST)
    if len(password) < 6:
        return Response({'detail': 'Password must be at least 6 characters.'}, status=status.HTTP_400_BAD_REQUEST)

    user = AuthUserModel.objects.create_user(username=username, password=password)
    user.is_staff = False
    user.is_superuser = False
    user.save()
    profile, _ = UserProfile.objects.get_or_create(user=user)
    profile.allowed_sections = sections
    profile.save()
    return Response(_serialize_managed_user(user), status=status.HTTP_201_CREATED)


@api_view(['PATCH', 'DELETE'])
@permission_classes([IsAuthenticated, IsAdminSection])
def admin_user_detail_view(request, pk):
    """PATCH: update sections / reset password. DELETE: remove the account."""
    try:
        user = AuthUserModel.objects.get(pk=pk)
    except AuthUserModel.DoesNotExist:
        return Response({'detail': 'User not found.'}, status=status.HTTP_404_NOT_FOUND)

    # Guard: never let an admin lock themselves out or edit an admin's access.
    if request.method == 'DELETE':
        if user.id == request.user.id:
            return Response({'detail': 'You cannot delete your own account.'}, status=status.HTTP_400_BAD_REQUEST)
        if is_admin_user(user):
            return Response({'detail': 'Admin accounts cannot be deleted here.'}, status=status.HTTP_400_BAD_REQUEST)
        user.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    # --- PATCH ---
    if is_admin_user(user):
        return Response({'detail': 'Admin accounts always have full access and cannot be edited here.'}, status=status.HTTP_400_BAD_REQUEST)

    # Update sections if provided.
    if 'allowed_sections' in request.data:
        sections = _clean_section_input(request.data.get('allowed_sections'))
        if sections is None:
            return Response({'detail': 'allowed_sections must be a list.'}, status=status.HTTP_400_BAD_REQUEST)
        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.allowed_sections = sections
        profile.save()
        # Once an explicit profile is set, drop the legacy group so it can't
        # override the per-user sections.
        user.groups.remove(*user.groups.filter(name=PNL_ONLY_GROUP))

    # Reset password if provided.
    new_password = request.data.get('password')
    if new_password:
        if len(new_password) < 6:
            return Response({'detail': 'Password must be at least 6 characters.'}, status=status.HTTP_400_BAD_REQUEST)
        user.set_password(new_password)
        user.save()
        # Invalidate existing tokens so the old session can't linger.
        Token.objects.filter(user=user).delete()

    # Toggle active state if provided.
    if 'is_active' in request.data:
        user.is_active = bool(request.data.get('is_active'))
        user.save()
        if not user.is_active:
            Token.objects.filter(user=user).delete()

    user.refresh_from_db()
    return Response(_serialize_managed_user(user))


# ---------------------------------------------------------------------------
# Billing Reminders
# ---------------------------------------------------------------------------
@api_view(['GET'])
@permission_classes([IsAuthenticated, BlockPnlOnly])
def billing_reminders_list(request):
    """List all billing reminders."""
    reminders = BillingReminder.objects.all()
    serializer = BillingReminderSerializer(reminders, many=True)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated, BlockPnlOnly])
def billing_reminder_create(request):
    """Create a new billing reminder."""
    serializer = BillingReminderSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
@permission_classes([IsAuthenticated, BlockPnlOnly])
def billing_reminder_update(request, pk):
    """Update a billing reminder."""
    try:
        reminder = BillingReminder.objects.get(pk=pk)
    except BillingReminder.DoesNotExist:
        return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

    serializer = BillingReminderSerializer(reminder, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated, BlockPnlOnly])
def billing_reminder_toggle_paid(request, pk):
    """Toggle the is_paid status of a billing reminder."""
    try:
        reminder = BillingReminder.objects.get(pk=pk)
    except BillingReminder.DoesNotExist:
        return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

    reminder.is_paid = not reminder.is_paid
    reminder.save()
    serializer = BillingReminderSerializer(reminder)
    return Response(serializer.data)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated, BlockPnlOnly])
def billing_reminder_delete(request, pk):
    """Delete a billing reminder."""
    try:
        reminder = BillingReminder.objects.get(pk=pk)
    except BillingReminder.DoesNotExist:
        return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

    reminder.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['POST'])
@permission_classes([IsAuthenticated, BlockPnlOnly])
def import_expenses(request):
    """Import expenses from Excel or CSV file."""
    file_obj = request.FILES.get('file')
    if not file_obj:
        return Response({'detail': 'No file was uploaded.'}, status=status.HTTP_400_BAD_REQUEST)

    import openpyxl
    from io import BytesIO

    try:
        fname = file_obj.name.lower()
        if fname.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        elif fname.endswith('.csv') or fname.endswith('.txt'):
            import csv
            file_data = file_obj.read().decode('utf-8-sig').splitlines()
            reader = csv.reader(file_data)
            rows = list(reader)
        else:
            return Response({'detail': 'Only .xlsx and .csv file formats are supported.'}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({'detail': f'Error reading file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

    if not rows:
        return Response({'detail': 'The uploaded file is empty.'}, status=status.HTTP_400_BAD_REQUEST)

    # Dynamically find the header row by looking for known keywords
    header_keywords = {'date', 'category', 'branch', 'amount', 'credit', 'debit', 'type', 'remark'}
    header_idx = 0
    headers = []

    for i, row in enumerate(rows):
        curr_headers = [str(cell).strip().lower() if cell is not None else '' for cell in row]
        # If this row has multiple header keywords, it's probably our header row
        if any(k in curr_headers or any(k in h for h in curr_headers) for k in header_keywords):
            headers = curr_headers
            header_idx = i
            break
    else:
        # Fallback to first row if keywords aren't matched explicitly
        headers = [str(cell).strip().lower() if cell is not None else '' for cell in rows[0]]
        header_idx = 0

    print(f"[IMPORT DEBUG] Found headers at row {header_idx + 1}: {headers}")
    print(f"[IMPORT DEBUG] Total data rows to process: {len(rows) - header_idx - 1}")
    
    header_mapping = {
        'date': 'date',
        'category': 'category',
        'branch': 'branch',
        'credit': 'credited_amount',
        'debit': 'debited_amount',
        'credit amount': 'credited_amount',
        'credit_amount': 'credited_amount',
        'credited_amount': 'credited_amount',
        'credit remark': 'credit_remark',
        'credit_remark': 'credit_remark',
        'credit person': 'credit_person',
        'credit_person': 'credit_person',
        'credit payment mode': 'credit_payment_mode',
        'credit_payment_mode': 'credit_payment_mode',
        'debit amount': 'debited_amount',
        'debit_amount': 'debited_amount',
        'debited_amount': 'debited_amount',
        'debit remark': 'debit_remark',
        'debit_remark': 'debit_remark',
        'debit person': 'debit_person',
        'debit_person': 'debit_person',
        'debit payment mode': 'debit_payment_mode',
        'debit_payment_mode': 'debit_payment_mode',
        'remark': 'remark',
        'person': 'person',
        'mode': 'mode',
        'payment mode': 'mode',
        'amount': 'amount',
        'type': 'type',
        'expense type': 'type',
        'transaction type': 'type',
    }
 
    # Load user-configured payment modes so custom/renamed modes (e.g. "IDFC Bank",
    # "Bank of Baroda") are preserved on import instead of collapsing to "Other".
    configured_modes = {
        m.strip().lower(): m
        for m in PaymentModeBalance.objects.values_list('payment_mode', flat=True)
        if m and m.strip()
    }

    import_data = []
    # Start processing data rows occurring AFTER the header row
    for row_idx, row in enumerate(rows[header_idx + 1:], header_idx + 2):
        if not any(cell is not None and str(cell).strip() != '' for cell in row):
            continue
 
        row_dict = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers):
                header = headers[col_idx]
                field = header_mapping.get(header)
                if field:
                    if cell is None or str(cell).strip() == '':
                        if field in ['credited_amount', 'debited_amount', 'amount']:
                            row_dict[field] = None
                        else:
                            row_dict[field] = ''
                    elif field == 'date':
                        import datetime as dt
                        if isinstance(cell, (dt.datetime, dt.date)):
                            if isinstance(cell, dt.datetime):
                                row_dict[field] = cell.date().isoformat()
                            else:
                                row_dict[field] = cell.isoformat()
                        else:
                            try:
                                date_str = str(cell).strip().split(' ')[0]
                                parsed_date = datetime.strptime(date_str, '%Y-%m-%d')
                                row_dict[field] = parsed_date.date().isoformat()
                            except ValueError:
                                try:
                                    date_str = str(cell).strip().split(' ')[0]
                                    parsed_date = datetime.strptime(date_str, '%d-%m-%Y')
                                    row_dict[field] = parsed_date.date().isoformat()
                                except ValueError:
                                    try:
                                        date_str = str(cell).strip().split(' ')[0]
                                        parsed_date = datetime.strptime(date_str, '%d/%m/%Y')
                                        row_dict[field] = parsed_date.date().isoformat()
                                    except ValueError:
                                        try:
                                            date_str = str(cell).strip().split(' ')[0]
                                            parsed_date = datetime.strptime(date_str, '%Y/%m/%d')
                                            row_dict[field] = parsed_date.date().isoformat()
                                        except ValueError:
                                            try:
                                                date_str = str(cell).strip().split(' ')[0]
                                                parsed_date = datetime.strptime(date_str, '%d.%m.%Y')
                                                row_dict[field] = parsed_date.date().isoformat()
                                            except ValueError:
                                                try:
                                                    date_str = str(cell).strip().split(' ')[0]
                                                    parsed_date = datetime.strptime(date_str, '%Y.%m.%d')
                                                    row_dict[field] = parsed_date.date().isoformat()
                                                except ValueError:
                                                    try:
                                                        date_str = str(cell).strip().split(' ')[0]
                                                        parsed_date = datetime.strptime(date_str, '%d-%b-%Y')
                                                        row_dict[field] = parsed_date.date().isoformat()
                                                    except ValueError:
                                                        try:
                                                            date_str = " ".join(str(cell).strip().split(' ')[:3])
                                                            parsed_date = datetime.strptime(date_str, '%d %b %Y')
                                                            row_dict[field] = parsed_date.date().isoformat()
                                                        except ValueError:
                                                            row_dict[field] = str(cell).strip()
                    elif field in ['credited_amount', 'debited_amount']:
                        try:
                            cleaned_val = str(cell).replace('₹', '').replace('Rs', '').replace(',', '').strip()
                            val = float(cleaned_val)
                            row_dict[field] = val if val > 0 else None
                        except (ValueError, TypeError):
                            row_dict[field] = None
                    elif field == 'amount':
                        try:
                            cleaned_val = str(cell).replace('₹', '').replace('Rs', '').replace(',', '').strip()
                            row_dict['amount'] = float(cleaned_val)
                        except (ValueError, TypeError):
                            row_dict['amount'] = None
                    elif field == 'type':
                        row_dict['type'] = str(cell).strip().lower()
                    else:
                        row_dict[field] = str(cell).strip()

        # Handle single 'amount' and 'type' column format if separate columns aren't filled
        if row_dict.get('amount') is not None:
            amt = row_dict['amount']
            t = row_dict.get('type', 'debit')
            if 'credit' in t:
                row_dict['credited_amount'] = amt
                row_dict['debited_amount'] = None
            else:
                row_dict['debited_amount'] = amt
                row_dict['credited_amount'] = None

        if 'credited_amount' not in row_dict:
            row_dict['credited_amount'] = None
        if 'debited_amount' not in row_dict:
            row_dict['debited_amount'] = None

        is_credit = row_dict.get('credited_amount') is not None

        if 'remark' in row_dict:
            if is_credit:
                row_dict['credit_remark'] = row_dict['remark']
                row_dict['debit_remark'] = ''
            else:
                row_dict['debit_remark'] = row_dict['remark']
                row_dict['credit_remark'] = ''

        if 'person' in row_dict:
            if is_credit:
                row_dict['credit_person'] = row_dict['person']
                row_dict['debit_person'] = ''
            else:
                row_dict['debit_person'] = row_dict['person']
                row_dict['credit_person'] = ''

        if 'mode' in row_dict:
            if is_credit:
                row_dict['credit_payment_mode'] = row_dict['mode']
                row_dict['debit_payment_mode'] = ''
            else:
                row_dict['debit_payment_mode'] = row_dict['mode']
                row_dict['credit_payment_mode'] = ''

        # Normalize payment modes. Preference order:
        #   1. Match a user-configured payment mode (case-insensitive) -> keep its stored casing
        #   2. Map well-known aliases to a canonical label
        #   3. Otherwise KEEP the original value (do NOT collapse to "Other")
        def normalize_mode(val):
            if not val:
                return ''
            raw = str(val).strip()
            val_lower = raw.lower()
            # 1. Preserve any mode the user has configured (e.g. "IDFC Bank", "Bank of Baroda")
            if val_lower in configured_modes:
                return configured_modes[val_lower]
            # 2. Known aliases
            if val_lower in ['cash', 'c']:
                return 'Cash'
            if val_lower in ['bank transfer', 'bank_transfer', 'bank', 'transfer']:
                return 'Bank Transfer'
            if val_lower in ['gpay', 'g-pay', 'google pay']:
                return 'GPay'
            if val_lower in ['phonepe', 'phone-pe', 'phone pe']:
                return 'PhonePe'
            if val_lower in ['upi']:
                return 'UPI'
            if val_lower in ['cheque']:
                return 'Cheque'
            # 3. Unknown but non-empty -> keep exactly what the sheet had
            return raw

        if row_dict.get('credit_payment_mode'):
            row_dict['credit_payment_mode'] = normalize_mode(row_dict['credit_payment_mode'])
        if row_dict.get('debit_payment_mode'):
            row_dict['debit_payment_mode'] = normalize_mode(row_dict['debit_payment_mode'])

        if 'branch' not in row_dict or not row_dict['branch']:
            row_dict['branch'] = 'Main Branch'
        if 'category' not in row_dict or not row_dict['category']:
            row_dict['category'] = 'Misc'
        if 'date' not in row_dict or not row_dict['date']:
            import datetime as dt
            row_dict['date'] = dt.date.today().isoformat()

        # Final sanitization: convert any leftover None/missing for string fields to empty string
        string_fields = [
            'credit_remark', 'debit_remark',
            'credit_person', 'debit_person',
            'credit_payment_mode', 'debit_payment_mode',
            'category', 'branch'
        ]
        for f in string_fields:
            if row_dict.get(f) is None:
                row_dict[f] = ''

        import_data.append((row_idx, row_dict))

    errors = []
    success_count = 0

    # Fields accepted by the serializer
    valid_fields = {
        'date', 'category', 'branch',
        'credited_amount', 'credit_remark', 'credit_person', 'credit_payment_mode',
        'debited_amount', 'debit_remark', 'debit_person', 'debit_payment_mode',
    }

    for row_idx, data in import_data:
        # Strip out extra keys that aren't in the serializer
        clean_data = {k: v for k, v in data.items() if k in valid_fields}
        print(f"[IMPORT DEBUG] Row {row_idx} clean_data: {clean_data}")
        serializer = ExpenseCreateSerializer(data=clean_data)
        if serializer.is_valid():
            serializer.save()
            success_count += 1
            print(f"[IMPORT DEBUG] Row {row_idx}: SAVED OK")
        else:
            err_msg = ", ".join([f"{k}: {', '.join(v)}" for k, v in serializer.errors.items()])
            errors.append(f"Row {row_idx}: {err_msg}")
            print(f"[IMPORT DEBUG] Row {row_idx} ERRORS: {serializer.errors}")

    if errors:
        return Response({
            'detail': f'Import completed. Successfully imported {success_count} of {success_count + len(errors)} entries.',
            'errors': errors,
            'success_count': success_count
        }, status=status.HTTP_200_OK)

    return Response({
        'detail': f'Successfully imported {success_count} expenses.',
        'success_count': success_count
    }, status=status.HTTP_201_CREATED)


class PettyCashDebitViewSet(viewsets.ModelViewSet):
    queryset = PettyCashDebit.objects.select_related('branch').all()
    serializer_class = PettyCashDebitSerializer
    pagination_class = None
    permission_classes = [IsAuthenticated, RequireAnySection(SECTION_EXPENSES, SECTION_PETTYCASH)]

    def get_queryset(self):
        qs = PettyCashDebit.objects.select_related('branch').all()
        branch_val = self.request.query_params.get('branch')
        if branch_val:
            if branch_val.isdigit():
                qs = qs.filter(branch_id=branch_val)
            else:
                qs = qs.filter(branch__location__icontains=branch_val)
        
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
            
        return qs.order_by('-date', '-created_at')


@api_view(['GET'])
@permission_classes([IsAuthenticated, RequireAnySection(SECTION_EXPENSES, SECTION_PETTYCASH)])
def petty_cash_summary(request):
    """Get petty cash summary, including credits (from Expenses) and debits."""
    credits_qs = Expense.objects.filter(category__icontains='petty')
    debits_qs = PettyCashDebit.objects.all()

    # Apply branch and date filters
    branch_val = request.query_params.get('branch')
    if branch_val:
        if branch_val.isdigit():
            credits_qs = credits_qs.filter(branch_id=branch_val)
            debits_qs = debits_qs.filter(branch_id=branch_val)
        else:
            credits_qs = credits_qs.filter(branch__location__icontains=branch_val)
            debits_qs = debits_qs.filter(branch__location__icontains=branch_val)

    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')
    if date_from:
        credits_qs = credits_qs.filter(date__gte=date_from)
        debits_qs = debits_qs.filter(date__gte=date_from)
    if date_to:
        credits_qs = credits_qs.filter(date__lte=date_to)
        debits_qs = debits_qs.filter(date__lte=date_to)

    # Calculate totals
    totals = credits_qs.aggregate(
        credits_sum=Coalesce(Sum('credited_amount'), Decimal('0.00')),
        debits_sum=Coalesce(Sum('debited_amount'), Decimal('0.00')),
    )
    total_credits = totals['credits_sum'] + totals['debits_sum']
    
    total_debits = debits_qs.aggregate(total=Coalesce(Sum('amount'), Decimal('0.00')))['total']
    balance = total_credits - total_debits

    credits_data = ExpenseSerializer(credits_qs.order_by('-date', '-created_at'), many=True).data
    debits_data = PettyCashDebitSerializer(debits_qs.order_by('-date', '-created_at'), many=True).data

    return Response({
        'balance': str(balance),
        'total_credits': str(total_credits),
        'total_debits': str(total_debits),
        'credits': credits_data,
        'debits': debits_data,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, RequireAnySection(SECTION_EXPENSES, SECTION_PETTYCASH)])
def export_petty_cash(request):
    """Export the petty cash ledger (credits from petty expenses + cash debits)
    as Excel or CSV, honouring the branch / date-range filters. Each row carries
    a running balance and a TOTAL line is appended."""
    credits_qs = Expense.objects.filter(category__icontains='petty').select_related('branch')
    debits_qs = PettyCashDebit.objects.select_related('branch').all()

    branch_val = request.query_params.get('branch')
    if branch_val:
        if branch_val.isdigit():
            credits_qs = credits_qs.filter(branch_id=branch_val)
            debits_qs = debits_qs.filter(branch_id=branch_val)
        else:
            credits_qs = credits_qs.filter(branch__location__icontains=branch_val)
            debits_qs = debits_qs.filter(branch__location__icontains=branch_val)
    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')
    if date_from:
        credits_qs = credits_qs.filter(date__gte=date_from)
        debits_qs = debits_qs.filter(date__gte=date_from)
    if date_to:
        credits_qs = credits_qs.filter(date__lte=date_to)
        debits_qs = debits_qs.filter(date__lte=date_to)

    # Merge credits and debits into one chronological ledger.
    rows = []
    for e in credits_qs:
        amt = (e.credited_amount or Decimal('0.00')) + (e.debited_amount or Decimal('0.00'))
        rows.append({'date': e.date, 'type': 'Credit', 'amount': amt,
                     'remark': e.credit_remark or e.debit_remark or '',
                     'person': e.credit_person or e.debit_person or '',
                     'branch': e.branch.location if e.branch else ''})
    for d in debits_qs:
        rows.append({'date': d.date, 'type': 'Debit', 'amount': d.amount or Decimal('0.00'),
                     'remark': d.remark or '', 'person': d.person or '',
                     'branch': d.branch.location if d.branch else ''})
    rows.sort(key=lambda r: (r['date'], 0 if r['type'] == 'Credit' else 1))

    running = Decimal('0.00')
    ledger = []
    for i, r in enumerate(rows, 1):
        running += r['amount'] if r['type'] == 'Credit' else -r['amount']
        ledger.append((i, r, running))

    total_credits = sum((r['amount'] for r in rows if r['type'] == 'Credit'), Decimal('0.00'))
    total_debits = sum((r['amount'] for r in rows if r['type'] == 'Debit'), Decimal('0.00'))
    balance = total_credits - total_debits

    headers = ['S.No', 'Date', 'Type', 'Credit (In)', 'Debit (Out)', 'Remark', 'Person', 'Branch', 'Balance']

    def row_values(i, r, run):
        credit_in = float(r['amount']) if r['type'] == 'Credit' else ''
        debit_out = float(r['amount']) if r['type'] == 'Debit' else ''
        return [i, r['date'].strftime('%Y-%m-%d'), r['type'], credit_in, debit_out,
                r['remark'], r['person'], r['branch'], float(run)]

    total_row = ['', '', 'TOTAL', float(total_credits), float(total_debits), '', '', '', float(balance)]

    if request.query_params.get('type', 'excel') == 'excel':
        try:
            import openpyxl
        except ImportError:
            return Response({'error': 'openpyxl not installed for Excel export'},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Petty Cash'
        ws.append(headers)
        for i, r, run in ledger:
            ws.append(row_values(i, r, run))
        ws.append([])
        ws.append(total_row)
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="petty-cash.xlsx"'
        return resp

    resp = HttpResponse(content_type='text/csv')
    resp['Content-Disposition'] = 'attachment; filename="petty-cash.csv"'
    writer = csv.writer(resp)
    writer.writerow(headers)
    for i, r, run in ledger:
        writer.writerow(row_values(i, r, run))
    writer.writerow([])
    writer.writerow(total_row)
    return resp


# ---------------------------------------------------------------------------
# Invoices
# ---------------------------------------------------------------------------
class InvoiceViewSet(viewsets.ModelViewSet):
    """CRUD for customer invoices (Tax Invoice / Bill of Supply)."""
    queryset = Invoice.objects.prefetch_related('items').all()
    serializer_class = InvoiceSerializer
    pagination_class = None
    permission_classes = [IsAuthenticated, RequireSection(SECTION_INVOICE)]

    def get_queryset(self):
        qs = Invoice.objects.prefetch_related('items').all()
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(invoice_number__icontains=search) |
                Q(customer_name__icontains=search)
            )
        return qs.order_by('-issue_date', '-created_at')


class DeliveryChallanViewSet(viewsets.ModelViewSet):
    """CRUD for delivery challans (goods delivery, no amounts)."""
    queryset = DeliveryChallan.objects.prefetch_related('items').all()
    serializer_class = DeliveryChallanSerializer
    pagination_class = None
    permission_classes = [IsAuthenticated, RequireSection(SECTION_CHALLAN)]

    def get_queryset(self):
        qs = DeliveryChallan.objects.prefetch_related('items').all()
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(challan_number__icontains=search) |
                Q(customer_name__icontains=search)
            )
        return qs.order_by('-challan_date', '-created_at')


class PurchaseBillViewSet(viewsets.ModelViewSet):
    """CRUD for purchase bills (goods/services bought from vendors)."""
    queryset = PurchaseBill.objects.prefetch_related('items').all()
    serializer_class = PurchaseBillSerializer
    pagination_class = None
    permission_classes = [IsAuthenticated, RequireSection(SECTION_PURCHASE)]

    def get_queryset(self):
        qs = PurchaseBill.objects.prefetch_related('items').all()
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(bill_number__icontains=search) |
                Q(vendor_name__icontains=search)
            )
        return qs.order_by('-issue_date', '-created_at')


class PurchaseOrderViewSet(viewsets.ModelViewSet):
    """CRUD for purchase orders (issued to vendors before buying)."""
    queryset = PurchaseOrder.objects.prefetch_related('items').all()
    serializer_class = PurchaseOrderSerializer
    pagination_class = None
    permission_classes = [IsAuthenticated, RequireSection(SECTION_PORDER)]

    def get_queryset(self):
        qs = PurchaseOrder.objects.prefetch_related('items').all()
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(order_number__icontains=search) |
                Q(vendor_name__icontains=search)
            )
        return qs.order_by('-issue_date', '-created_at')


class PaymentReceiptViewSet(viewsets.ModelViewSet):
    """CRUD for payment receipts (money received from customers)."""
    queryset = PaymentReceipt.objects.prefetch_related('lines').all()
    serializer_class = PaymentReceiptSerializer
    pagination_class = None
    permission_classes = [IsAuthenticated, RequireSection(SECTION_RECEIPT)]

    def get_queryset(self):
        qs = PaymentReceipt.objects.prefetch_related('lines').all()
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(receipt_number__icontains=search) |
                Q(receipt_to_name__icontains=search)
            )
        return qs.order_by('-payment_date', '-created_at')


class QuoteViewSet(viewsets.ModelViewSet):
    """CRUD for price quotes issued to customers."""
    queryset = Quote.objects.prefetch_related('items').all()
    serializer_class = QuoteSerializer
    pagination_class = None
    permission_classes = [IsAuthenticated, RequireSection(SECTION_QUOTE)]

    def get_queryset(self):
        qs = Quote.objects.prefetch_related('items').all()
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(quote_number__icontains=search) |
                Q(customer_name__icontains=search)
            )
        return qs.order_by('-issue_date', '-created_at')


class BillOfSupplyViewSet(viewsets.ModelViewSet):
    """CRUD for bills of supply (sales without GST)."""
    queryset = BillOfSupply.objects.prefetch_related('items').all()
    serializer_class = BillOfSupplySerializer
    pagination_class = None
    permission_classes = [IsAuthenticated, RequireSection(SECTION_BOS)]

    def get_queryset(self):
        qs = BillOfSupply.objects.prefetch_related('items').all()
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(Q(bos_number__icontains=search) | Q(customer_name__icontains=search))
        return qs.order_by('-issue_date', '-created_at')


class TaxInvoiceViewSet(viewsets.ModelViewSet):
    """CRUD for GST tax invoices (CGST+SGST intra-state / IGST inter-state)."""
    queryset = TaxInvoice.objects.prefetch_related('items').all()
    serializer_class = TaxInvoiceSerializer
    pagination_class = None
    permission_classes = [IsAuthenticated, RequireSection(SECTION_TAXINVOICE)]

    def get_queryset(self):
        qs = TaxInvoice.objects.prefetch_related('items').all()
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(Q(ti_number__icontains=search) | Q(customer_name__icontains=search))
        return qs.order_by('-issue_date', '-created_at')


# ---------------------------------------------------------------------------
# Bank statement import (IDFC FIRST Bank / BOB)
# ---------------------------------------------------------------------------
# Users upload the bank's own Excel/CSV export; we auto-detect the header row
# and map the (bank-specific) column names to canonical fields, tolerating the
# many header spellings IDFC FIRST Bank and Bank of Baroda use.

_BANK_MAX_ROWS = 50000


def _smart_decode(raw):
    """Decode file bytes to text, tolerating the various encodings banks emit."""
    if raw[:2] in (b'\xff\xfe', b'\xfe\xff'):
        try:
            return raw.decode('utf-16')
        except Exception:
            pass
    for enc in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
        try:
            return raw.decode(enc)
        except Exception:
            continue
    return raw.decode('utf-8', errors='replace')


def _parse_html_table(raw):
    """Extract table rows from an HTML file OR a Microsoft XML-Spreadsheet file
    (the 'Excel' many Indian bank portals actually export). Uses only the stdlib
    html parser, so no extra dependency is required. Returns a list of row lists
    (all tables/sheets flattened; the caller re-detects the real header row)."""
    from html.parser import HTMLParser

    text = _smart_decode(raw)

    class _TableExtractor(HTMLParser):
        def __init__(self):
            super().__init__(convert_charrefs=True)
            self.rows = []
            self.cur = None
            self.buf = []
            self.in_cell = False

        def _base(self, tag):
            return tag.lower().split(':')[-1]  # handles namespaced ss:Row / ss:Cell

        def handle_starttag(self, tag, attrs):
            base = self._base(tag)
            if base in ('tr', 'row'):
                self.cur = []
            elif base in ('td', 'th', 'cell'):
                self.in_cell = True
                self.buf = []
            elif base == 'br' and self.in_cell:
                self.buf.append(' ')

        def handle_startendtag(self, tag, attrs):
            base = self._base(tag)
            if base in ('td', 'th', 'cell') and self.cur is not None:
                self.cur.append('')

        def handle_endtag(self, tag):
            base = self._base(tag)
            if base in ('td', 'th', 'cell') and self.cur is not None and self.in_cell:
                self.cur.append(' '.join(''.join(self.buf).split()))
                self.in_cell = False
            elif base in ('tr', 'row') and self.cur is not None:
                self.rows.append(self.cur)
                self.cur = None

        def handle_data(self, data):
            if self.in_cell:
                self.buf.append(data)

    p = _TableExtractor()
    p.feed(text)
    return [r for r in p.rows if any((str(c) or '').strip() for c in r)]


def _read_tabular(file_obj):
    """Read an uploaded bank export into a list of rows (each a list of cells).

    Handles, in order: real .csv/.txt, real .xlsx (zip), HTML tables saved as
    .xls/.xlsx (common with Indian bank portals), MS XML-Spreadsheet, and old
    binary .xls (BIFF, if xlrd is available). Raises ValueError with a friendly
    message otherwise."""
    from io import BytesIO
    fname = (file_obj.name or '').lower()
    raw = file_obj.read()
    if not raw:
        raise ValueError('The uploaded file is empty.')

    # 1) Plain CSV / TXT by extension.
    if fname.endswith('.csv') or fname.endswith('.txt'):
        return [row for row in csv.reader(_smart_decode(raw).splitlines())]

    # 2) Real .xlsx (an OOXML zip). Try this first regardless of extension.
    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(raw), data_only=True, read_only=True)
        ws = wb.active
        return [list(r) for r in ws.iter_rows(values_only=True)]
    except Exception:
        pass  # not a real .xlsx — fall through to the sniffers below

    head = raw[:8192].lstrip().lower()

    # 3) HTML table or MS XML-Spreadsheet ("Excel" that's really markup).
    if (head[:1] == b'<' or b'<table' in head or b'<html' in head
            or b'<tr' in head or b'spreadsheet' in head or b'<?xml' in head):
        rows = _parse_html_table(raw)
        if rows:
            return rows

    # 4) Old binary .xls (BIFF / OLE2), read with xlrd. Date-typed cells come
    # back as Excel serial floats, so convert them to real datetimes here.
    if raw[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        try:
            import xlrd
            book = xlrd.open_workbook(file_contents=raw)
            sh = book.sheet_by_index(0)
            out = []
            for r in range(sh.nrows):
                row = []
                for c in range(sh.ncols):
                    cell = sh.cell(r, c)
                    val = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            val = xlrd.xldate.xldate_as_datetime(val, book.datemode)
                        except Exception:
                            pass
                    row.append(val)
                out.append(row)
            return out
        except ImportError:
            raise ValueError("This is an old .xls file. Open it in Excel and use "
                             "'Save As' → 'Excel Workbook (.xlsx)' or 'CSV', then upload again.")
        except Exception:
            raise ValueError('Could not read this .xls file. Please re-save it as .xlsx or CSV and try again.')

    # 5) Last resort: maybe it's delimited text with the wrong extension.
    try:
        text = _smart_decode(raw)
        if ',' in text or '\t' in text:
            rows = [row for row in csv.reader(text.splitlines())]
            if rows:
                return rows
    except Exception:
        pass

    raise ValueError("Could not read the file. Please upload the bank's Excel (.xlsx) "
                     "or CSV export. If it's a .xls, re-save it as .xlsx first.")


def _match_bank_header(h):
    """Map one raw header cell to a canonical bank-statement field, or None."""
    h = str(h or '').strip().lower()
    if not h:
        return None
    contains = lambda *subs: any(s in h for s in subs)
    # Value date must be checked before the generic 'date' rule below.
    if contains('value date', 'value dt'):
        return 'value_date'
    if contains('withdrawal', 'debit', 'paid out', 'amount debited') or h in ('dr', 'withdrawals', 'dr amount'):
        return 'debit'
    if contains('deposit', 'credit', 'paid in', 'amount credited') or h in ('cr', 'deposits', 'cr amount'):
        return 'credit'
    if contains('closing balance', 'running balance') or h == 'balance' or contains('balance'):
        return 'balance'
    if contains('narration', 'particular', 'description', 'remark', 'details', 'transaction detail'):
        return 'narration'
    if contains('chq', 'cheque', 'ref no', 'reference', 'instrument', 'utr'):
        return 'ref_no'
    if contains('date', 'txn dt', 'tran date', 'posting'):
        return 'txn_date'
    return None


def _parse_bank_date(cell):
    import datetime as dt
    if cell is None or str(cell).strip() == '':
        return None
    if isinstance(cell, dt.datetime):
        return cell.date()
    if isinstance(cell, dt.date):
        return cell
    s = str(cell).strip().split(' ')[0]
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y', '%d-%m-%y', '%Y-%m-%d',
                '%d-%b-%Y', '%d-%b-%y', '%d %b %Y', '%d/%b/%Y', '%d.%m.%Y', '%m/%d/%Y'):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_bank_amount(cell):
    """Parse an amount cell → Decimal. Blank/'-'/None → 0. Tolerates commas,
    currency prefixes, Cr/Dr suffixes and (parenthesised) negatives."""
    if cell is None:
        return Decimal('0')
    if isinstance(cell, (int, float)):
        return Decimal(str(cell))
    s = str(cell).strip()
    if s == '' or s in ('-', '.', 'nil', 'NIL', 'Nil'):
        return Decimal('0')
    neg = False
    if s.startswith('(') and s.endswith(')'):
        neg, s = True, s[1:-1]
    su = s.upper()
    for junk in ('CR', 'DR', 'INR', 'RS.', 'RS', '₹', ',', ' '):
        su = su.replace(junk, '')
    try:
        from decimal import InvalidOperation
        val = Decimal(su)
    except Exception:
        return Decimal('0')
    return -val if neg else val


def _bank_norm_text(s):
    """Collapse whitespace + upper-case so trivial formatting differences in a
    bank's re-export (extra spaces, case) don't defeat duplicate detection."""
    return ' '.join(str(s or '').split()).upper()


def _bank_norm_amount(amount):
    """Canonical 2-decimal string for an amount so 100 / 100.0 / 100.00 all
    match. None (blank balance) → ''."""
    if amount is None:
        return ''
    try:
        return str(Decimal(amount).quantize(Decimal('0.01')))
    except Exception:
        return str(amount)


def _bank_fingerprint(bank, txn_date, narration, debit, credit, balance, ref_no):
    """Stable per-row fingerprint used for duplicate detection. Computed the SAME
    way for existing DB rows and freshly-parsed rows, so re-uploading a statement
    (even one that overlaps earlier dates, or was re-exported with slightly
    different spacing) never creates duplicates."""
    return (
        f"{bank}|{txn_date}|{_bank_norm_text(narration)}|"
        f"{_bank_norm_amount(debit)}|{_bank_norm_amount(credit)}|"
        f"{_bank_norm_amount(balance)}|{_bank_norm_text(ref_no)}"
    )


def _import_bank_rows(rows, bank, source_file):
    """Parse already-read rows into BankStatementEntry objects for `bank`.
    Returns {inserted, skipped, errors}. Duplicate rows (same fingerprint) are
    skipped so re-uploading the same statement is safe."""
    import hashlib
    result = {'inserted': 0, 'skipped': 0, 'errors': []}
    if not rows:
        result['errors'].append('The file is empty.')
        return result

    # 1) Locate the header row. The real transaction header always has a DATE
    # column plus at least one money column; requiring the date lets us skip
    # summary blocks like "Opening Balance | Total Debit | Total Credit | ..."
    # that some banks (e.g. IDFC) print above the transaction table.
    header_idx, col_map = None, {}
    for i, row in enumerate(rows[:80]):
        mapping = {}
        for col_idx, cell in enumerate(row):
            field = _match_bank_header(cell)
            if field and field not in mapping.values():
                mapping[col_idx] = field
        vals = set(mapping.values())
        if len(mapping) >= 3 and 'txn_date' in vals and (vals & {'debit', 'credit', 'balance'}):
            header_idx, col_map = i, mapping
            break
    if header_idx is None:
        result['errors'].append('Could not find the statement columns (Date / Narration / Debit / Credit / Balance). Please upload the bank\'s original Excel export.')
        return result

    # 2) Pre-load existing fingerprints for this bank to skip duplicates.
    # Recompute each existing row's fingerprint from its FIELDS (not the stored
    # row_hash) so dedup stays consistent even for rows saved under an older hash
    # formula — this makes the very next import safe, with no data migration.
    seen = set()
    for e in BankStatementEntry.objects.filter(bank=bank).values(
        'txn_date', 'narration', 'debit', 'credit', 'balance', 'ref_no'
    ):
        seen.add(_bank_fingerprint(
            bank, e['txn_date'], e['narration'], e['debit'], e['credit'],
            e['balance'], e['ref_no'],
        ))
    to_create = []
    data_rows = rows[header_idx + 1:_BANK_MAX_ROWS + header_idx + 1]
    for row in data_rows:
        if not any(c is not None and str(c).strip() != '' for c in row):
            continue
        get = lambda field: next((row[ci] for ci, f in col_map.items() if f == field and ci < len(row)), None)
        txn_date = _parse_bank_date(get('txn_date'))
        debit = _parse_bank_amount(get('debit'))
        credit = _parse_bank_amount(get('credit'))
        narration = str(get('narration') or '').strip()
        # Every real transaction row carries a date; dateless rows are repeated
        # headers, blank separators or "Total"/summary footers — skip them.
        if txn_date is None:
            continue
        # A dated row with no money and no narration is a summary/blank footer
        # line (e.g. the closing-balance row), not a real transaction.
        if debit == 0 and credit == 0 and not narration:
            continue
        bal_raw = get('balance')
        balance = None if (bal_raw is None or str(bal_raw).strip() == '') else _parse_bank_amount(bal_raw)
        balance_dc = ''
        if bal_raw is not None:
            _bs = str(bal_raw).strip().upper()
            if _bs.endswith('CR'):
                balance_dc = 'Cr'
            elif _bs.endswith('DR'):
                balance_dc = 'Dr'
        ref_no = str(get('ref_no') or '').strip()[:150]
        value_date = _parse_bank_date(get('value_date'))

        fingerprint = _bank_fingerprint(bank, txn_date, narration, debit, credit, balance, ref_no)
        row_hash = hashlib.sha256(fingerprint.encode('utf-8')).hexdigest()
        if fingerprint in seen:
            result['skipped'] += 1
            continue
        seen.add(fingerprint)
        to_create.append(BankStatementEntry(
            bank=bank, txn_date=txn_date, value_date=value_date, narration=narration,
            ref_no=ref_no, debit=debit, credit=credit, balance=balance, balance_dc=balance_dc,
            source_file=source_file[:255], row_hash=row_hash,
        ))

    # Normalise stored order to oldest→newest so ascending id == chronological
    # order, whether the bank exported newest-first (BOB) or oldest-first (IDFC).
    # The list view then uses -id to surface the latest transaction on top and
    # to read the true current balance off the most recent row.
    dated = [e.txn_date for e in to_create if e.txn_date]
    if len(dated) >= 2 and dated[0] > dated[-1]:
        to_create.reverse()
    if to_create:
        BankStatementEntry.objects.bulk_create(to_create)
        result['inserted'] = len(to_create)
    elif result['skipped'] == 0:
        result['errors'].append('No transaction rows were found below the header.')
    return result


# Which expense payment-mode names belong to each bank. Fuzzy (substring, case-
# insensitive) so it survives the exact configured label — e.g. "IDFC Bank",
# "IDFC FIRST", "Bank of Baroda", "BOB Current" all map correctly.
_BANK_MODE_HINTS = {
    BankStatementEntry.BANK_IDFC: ('idfc',),
    BankStatementEntry.BANK_BOB: ('baroda', 'bob'),
}


def _bank_match_map(bank, entries):
    """Reconcile bank-statement `entries` against the Expenses ledger.

    A bank row counts as "matched" when the Expenses section holds an entry with
    the SAME date, SAME amount, on the SAME side (a bank debit ↔ an expense
    debit; a bank credit ↔ an expense credit) AND under a payment mode that
    belongs to this bank. Matching is one-to-one: each expense can satisfy only
    one bank row, so duplicate amounts on a day are reconciled by count.

    Returns ``(match_map, suggested_mode)`` where match_map is a dict
    ``{bank_entry_id: expense_detail}`` describing the matched Expenses entry (so
    the UI can show "what did I record for this transaction?" when the status is
    clicked), and suggested_mode is the payment-mode label the user most often
    uses for this bank (so a missing row can pre-fill the Add-Expense form).
    """
    from collections import defaultdict, Counter

    hints = _BANK_MODE_HINTS.get(bank, ())
    if not hints:
        return {}, ''

    def belongs(mode):
        m = (mode or '').lower()
        return any(h in m for h in hints)

    def q2(amount):
        try:
            return Decimal(amount).quantize(Decimal('0.01'))
        except Exception:
            return amount

    # Pull only the expenses whose mode could belong to this bank.
    mode_q = Q()
    for h in hints:
        mode_q |= Q(debit_payment_mode__icontains=h) | Q(credit_payment_mode__icontains=h)
    exp = Expense.objects.filter(mode_q).values(
        'id', 'date', 'category', 'branch__location',
        'debited_amount', 'credited_amount',
        'debit_payment_mode', 'credit_payment_mode',
        'debit_remark', 'credit_remark', 'debit_person', 'credit_person',
    )

    # Available expense "slots" keyed by (date, amount, side). Each key holds a
    # queue of the actual matching expense details, popped as bank rows claim them.
    slots = defaultdict(list)
    mode_counter = Counter()
    for e in exp:
        d = e['date']
        if e['debited_amount'] and belongs(e['debit_payment_mode']):
            mode_counter[e['debit_payment_mode']] += 1
            slots[(d, q2(e['debited_amount']), 'debit')].append({
                'id': e['id'], 'date': d.isoformat() if d else None,
                'category': e['category'], 'branch': e['branch__location'],
                'side': 'debit', 'amount': str(e['debited_amount']),
                'mode': e['debit_payment_mode'], 'remark': e['debit_remark'],
                'person': e['debit_person'],
            })
        if e['credited_amount'] and belongs(e['credit_payment_mode']):
            mode_counter[e['credit_payment_mode']] += 1
            slots[(d, q2(e['credited_amount']), 'credit')].append({
                'id': e['id'], 'date': d.isoformat() if d else None,
                'category': e['category'], 'branch': e['branch__location'],
                'side': 'credit', 'amount': str(e['credited_amount']),
                'mode': e['credit_payment_mode'], 'remark': e['credit_remark'],
                'person': e['credit_person'],
            })
    suggested_mode = mode_counter.most_common(1)[0][0] if mode_counter else ''

    # Claim slots in a stable order (oldest first) so the pairing is deterministic
    # regardless of the display ordering.
    match_map = {}
    for be in sorted(entries, key=lambda x: (x.txn_date or datetime.min.date(), x.id)):
        key = None
        if be.debit and be.debit > 0:
            key = (be.txn_date, q2(be.debit), 'debit')
        elif be.credit and be.credit > 0:
            key = (be.txn_date, q2(be.credit), 'credit')
        if key and slots.get(key):
            match_map[be.id] = slots[key].pop(0)
    return match_map, suggested_mode


def _bank_for_mode(mode):
    """Which bank (if any) an expense payment-mode label belongs to, else None."""
    m = (mode or '').lower()
    for bank, hints in _BANK_MODE_HINTS.items():
        if any(h in m for h in hints):
            return bank
    return None


def _bank_matched_expense_id_set(bank):
    """Set of Expense ids that reconcile to this bank's statement — the reverse
    view of _bank_match_map, used to flag expenses as in/out of the statement."""
    entries = list(BankStatementEntry.objects.filter(bank=bank))
    match_map, _ = _bank_match_map(bank, entries)
    return {d['id'] for d in match_map.values()}


def _bank_statement_by_expense(bank):
    """Reverse of _bank_match_map: returns ``{expense_id: statement_detail}`` so
    the Expenses UI can show the bank-statement row an entry reconciled to when
    its "In Statement" status is clicked."""
    entries = list(BankStatementEntry.objects.filter(bank=bank))
    by_id = {e.id: e for e in entries}
    match_map, _ = _bank_match_map(bank, entries)
    result = {}
    for be_id, exp_detail in match_map.items():
        be = by_id.get(be_id)
        if not be:
            continue
        result[exp_detail['id']] = {
            'id': be.id,
            'bank': bank,
            'bank_display': be.get_bank_display(),
            'txn_date': be.txn_date.isoformat() if be.txn_date else None,
            'value_date': be.value_date.isoformat() if be.value_date else None,
            'narration': be.narration,
            'ref_no': be.ref_no,
            'debit': str(be.debit),
            'credit': str(be.credit),
            'balance': str(be.balance) if be.balance is not None else None,
            'balance_dc': be.balance_dc,
        }
    return result


class _BankStatementViewSet(viewsets.ModelViewSet):
    """Base viewset for a single bank's statement entries. Subclasses set
    `bank` and the section permission."""
    serializer_class = BankStatementEntrySerializer
    pagination_class = None
    bank = None

    def get_queryset(self):
        qs = BankStatementEntry.objects.filter(bank=self.bank)
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(Q(narration__icontains=search) | Q(ref_no__icontains=search))
        # Optional transaction-date range filter (YYYY-MM-DD).
        date_from = self.request.query_params.get('from')
        date_to = self.request.query_params.get('to')
        if date_from:
            qs = qs.filter(txn_date__gte=date_from)
        if date_to:
            qs = qs.filter(txn_date__lte=date_to)
        # Most recent transactions first. Storage is normalised to oldest→newest
        # at import, so -id reliably surfaces the latest transaction of each day
        # on top (and lets the UI read the current balance off the first row).
        return qs.order_by('-txn_date', '-id')

    def list(self, request, *args, **kwargs):
        """List entries, annotating each with `expense_status` — whether the row
        has a matching entry in the Expenses ledger — and `matched_expense`, the
        details of that entry (or null) so the UI can show it on demand."""
        qs = self.filter_queryset(self.get_queryset())
        entries = list(qs)
        match_map, suggested_mode = _bank_match_map(self.bank, entries)
        data = self.get_serializer(entries, many=True).data
        for row in data:
            m = match_map.get(row['id'])
            row['expense_status'] = 'matched' if m else 'missing'
            row['matched_expense'] = m
        return Response({
            'results': data,
            'summary': {
                'total': len(entries),
                'matched': len(match_map),
                'missing': len(entries) - len(match_map),
                'suggested_mode': suggested_mode,
            },
        })

    @action(detail=False, methods=['post'], url_path='import')
    def import_statement(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'No file was uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            rows = _read_tabular(file_obj)
        except ValueError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        res = _import_bank_rows(rows, self.bank, file_obj.name)
        if res['inserted'] == 0 and res['errors']:
            return Response({'detail': res['errors'][0], **res}, status=status.HTTP_400_BAD_REQUEST)
        parts = [f"Imported {res['inserted']} entr{'y' if res['inserted'] == 1 else 'ies'}"]
        if res['skipped']:
            parts.append(f"skipped {res['skipped']} duplicate row(s)")
        res['detail'] = ', '.join(parts) + '.'
        return Response(res, status=status.HTTP_200_OK)

    @action(detail=False, methods=['delete'], url_path='clear')
    def clear(self, request):
        n, _ = BankStatementEntry.objects.filter(bank=self.bank).delete()
        return Response({'detail': f'Cleared {n} entries.', 'deleted': n})


class IDFCStatementViewSet(_BankStatementViewSet):
    """IDFC FIRST Bank statement entries + Excel import."""
    bank = BankStatementEntry.BANK_IDFC
    permission_classes = [IsAuthenticated, RequireSection(SECTION_IDFC)]


class BOBStatementViewSet(_BankStatementViewSet):
    """Bank of Baroda statement entries + Excel import."""
    bank = BankStatementEntry.BANK_BOB
    permission_classes = [IsAuthenticated, RequireSection(SECTION_BOB)]


# ---------------------------------------------------------------------------
# Engineer P&L — live profit/loss per engineer, closed-calls pulled from OpenCall
# ---------------------------------------------------------------------------
class EngineerPnlViewSet(viewsets.ModelViewSet):
    """CRUD for each engineer's P&L parameters, plus a live ``/board/`` action
    that pulls the closed-call count per engineer from the OpenCall system and
    computes Revenue / Nett in real time. The board degrades gracefully: if
    OpenCall is unreachable or unconfigured it still returns every engineer with
    closed = 0 and a ``live_ok: false`` flag, so the section never breaks."""
    queryset = EngineerPnl.objects.all()
    serializer_class = EngineerPnlSerializer
    pagination_class = None
    permission_classes = [IsAuthenticated, RequireSection(SECTION_ENGPNL)]

    def get_queryset(self):
        return EngineerPnl.objects.all().order_by('position', 'id')

    @action(detail=False, methods=['get'])
    def board(self, request):
        from datetime import date
        import calendar
        from . import opencall_client

        # Resolve the window. Priority: explicit from/to range → month → current
        # calendar month (the live default). from/to are YYYY-MM-DD.
        month = request.query_params.get('month')     # 'YYYY-MM'
        q_from = request.query_params.get('from')     # 'YYYY-MM-DD'
        q_to = request.query_params.get('to')
        try:
            if q_from and q_to:
                first = date.fromisoformat(q_from)
                last = date.fromisoformat(q_to)
                if first > last:
                    first, last = last, first
                year, mon = first.year, first.month
            elif month:
                year, mon = int(month[:4]), int(month[5:7])
                first = date(year, mon, 1)
                last = date(year, mon, calendar.monthrange(year, mon)[1])
            else:
                # Default view = today only (the live current-day board).
                today = date.today()
                first = last = today
                year, mon = today.year, today.month
        except (ValueError, IndexError):
            return Response({'detail': 'Invalid date range.'}, status=status.HTTP_400_BAD_REQUEST)

        period_days = (last - first).days + 1

        # Auto-populate engineers from OpenCall's roster so they appear without
        # manual entry. Non-destructive: only creates names never seen before
        # (a soft-hidden engineer — active=False — is NOT recreated). Best-effort.
        synced = 0
        if request.query_params.get('sync', '1') != '0':
            try:
                roster = opencall_client.get_engineers()
                known = {e.engineer_name.strip().lower() for e in EngineerPnl.objects.all()}
                to_add = []
                for oc in roster:
                    if oc['active'] and oc['name'].lower() not in known:
                        to_add.append(EngineerPnl(engineer_name=oc['name'], email=oc.get('email', '') or ''))
                        known.add(oc['name'].lower())
                if to_add:
                    EngineerPnl.objects.bulk_create(to_add)
                    synced = len(to_add)
            except Exception:
                pass  # roster sync is best-effort; never break the board

        engineers = list(self.get_queryset().filter(active=True))

        # Pull real salary from the Payroll system and apply it to matched
        # engineers (email first, then name). Payroll is the source of truth for
        # salary when a match exists; best-effort so it never breaks the board.
        payroll_ok, payroll_message = None, ''
        salary_source = {}  # engineer id -> 'payroll' | 'manual'
        try:
            from . import payroll_client
            if payroll_client.is_configured():
                sal = payroll_client.get_salaries()
                to_update = []
                for eng in engineers:
                    email_key = (eng.email or '').strip().lower()
                    name_key = eng.engineer_name.strip().lower()
                    match = sal['by_email'].get(email_key) if email_key else None
                    if match is None:
                        match = sal['by_name'].get(name_key)
                    if match is not None:
                        salary_source[eng.id] = 'payroll'
                        new_sal = Decimal(str(match)).quantize(Decimal('0.01'))
                        if eng.engg_salary != new_sal:
                            eng.engg_salary = new_sal
                            to_update.append(eng)
                if to_update:
                    EngineerPnl.objects.bulk_update(to_update, ['engg_salary'])
                payroll_ok = True
            else:
                payroll_ok = False
                payroll_message = 'Payroll credentials not set (PAYROLL_USERNAME / PAYROLL_PASSWORD).'
        except Exception as e:
            payroll_ok = False
            payroll_message = f'Could not reach Payroll: {e}'

        # Live pull of closed-call counts (name-keyed, case-insensitive).
        live_ok, message, counts, display, meta = True, '', {}, {}, {}
        try:
            res = opencall_client.get_closed_counts(first.isoformat(), last.isoformat())
            counts, display, meta = res['counts'], res['display'], res['meta']
        except opencall_client.OpenCallError as e:
            live_ok, message = False, str(e)
        except Exception as e:  # network/timeout etc. — never break the page
            live_ok, message = False, f'Could not reach OpenCall: {e}'

        configured_keys = {e.engineer_name.strip().lower() for e in engineers}
        all_rows = []
        for eng in engineers:
            closed = int(counts.get(eng.engineer_name.strip().lower(), 0))
            calc = eng.compute(closed, period_days)
            all_rows.append({
                'id': eng.id,
                'engineer_name': eng.engineer_name,
                'email': eng.email,
                'engg_count': eng.engg_count,
                'per_day_target': eng.per_day_target,
                'per_call_rate': str(eng.per_call_rate),
                'engg_salary': str(eng.engg_salary),
                'total_working_days': eng.total_working_days,
                'actual_working_days': eng.actual_working_days,
                'salary_source': salary_source.get(eng.id, 'manual'),
                **calc,
            })

        # Default = only engineers WITH data (closed calls) in the window; the
        # "Overall" view (?all=1) shows every configured engineer. When OpenCall
        # is offline we can't tell who has data, so fall back to showing all.
        show_all = request.query_params.get('all') == '1'
        rows = all_rows if (show_all or not live_ok) else [r for r in all_rows if r['closed_calls'] > 0]

        tot_engg = sum(r['engg_count'] for r in rows)
        tot_closed = sum(r['closed_calls'] for r in rows)
        tot_rev = sum((Decimal(r['revenue']) for r in rows), Decimal('0.00'))
        tot_sal = sum((Decimal(r['total_engg_salary']) for r in rows), Decimal('0.00'))
        tot_nett = sum((Decimal(r['nett']) for r in rows), Decimal('0.00'))

        # Engineers OpenCall reports that aren't configured here yet (so the user
        # can add them and start earning revenue for their closes).
        unmatched = sorted(
            [{'engineer_name': display[k], 'closed_calls': counts[k]}
             for k in counts if k not in configured_keys],
            key=lambda x: -x['closed_calls'],
        )

        return Response({
            'period': {'month': f'{year:04d}-{mon:02d}', 'from': first.isoformat(), 'to': last.isoformat()},
            'live_ok': live_ok,
            'message': message,
            'synced': synced,
            'payroll_ok': payroll_ok,
            'payroll_message': payroll_message,
            'show_all': show_all,
            'total_configured': len(all_rows),
            'meta': meta,
            'rows': rows,
            'totals': {
                'engg_count': tot_engg,
                'closed_calls': tot_closed,
                'revenue': str(tot_rev),
                'total_engg_salary': str(tot_sal),
                'nett': str(tot_nett),
            },
            'unmatched_engineers': unmatched,
        })


# ---------------------------------------------------------------------------
# Sleek Bill Invoice Register — import the Sleek Bill invoice export (.xls) and
# mirror its list (Tax Invoice + Bill of Supply) in the Invoice Register section.
# ---------------------------------------------------------------------------
def _parse_sleekbill_date(cell):
    import datetime as _dt
    if cell is None or str(cell).strip() == '':
        return None
    if isinstance(cell, (_dt.datetime, _dt.date)):
        return cell.date() if isinstance(cell, _dt.datetime) else cell
    s = str(cell).strip().split(' ')[0]
    for fmt in ('%d-%b-%Y', '%d-%B-%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%d-%b-%y', '%d/%b/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_sleekbill_amount(cell):
    if cell is None:
        return Decimal('0')
    s = str(cell).strip().replace(',', '').replace('₹', '')
    if s == '' or s in ('-', 'NA', 'N/A'):
        return Decimal('0')
    try:
        return Decimal(s)
    except Exception:
        return Decimal('0')


# Sleek Bill export column header (lowercased) -> model field.
_SB_COLUMN_MAP = {
    'client name': 'client_name', 'client gstin': 'client_gstin',
    'invoice number': 'invoice_number', 'creator name': 'creator_name',
    'client phone number': 'client_phone', 'client email': 'client_email',
    'client city': 'client_city', 'client state': 'client_state',
    'issue date': 'issue_date', 'due date': 'due_date', 'date of payment': 'date_of_payment',
    'payment mode': 'payment_mode', 'financial year': 'financial_year', 'currency': 'currency',
    'amount': 'amount', 'tax': 'tax', 'total': 'total',
    'status': 'status', 'amount paid': 'amount_paid', 'balance': 'balance',
    'dr. / cr.': 'dr_cr', 'type': 'invoice_type', 'payments': 'payment_info',
    'cgst': 'cgst', 'sgst': 'sgst', 'igst': 'igst',
}
_SB_DATE_FIELDS = {'issue_date', 'due_date', 'date_of_payment'}
_SB_AMOUNT_FIELDS = {'amount', 'tax', 'total', 'amount_paid', 'balance', 'cgst', 'sgst', 'igst'}


def _import_sleekbill_rows(rows, source_file):
    """Parse Sleek Bill export rows into SleekBillInvoice records (upsert by
    invoice_number). Returns {created, updated, skipped, errors}."""
    result = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': []}
    if not rows:
        result['errors'].append('The file is empty.')
        return result

    # Locate the header row (must contain an "invoice number" column).
    header_idx, col_map = None, {}
    for i, row in enumerate(rows[:20]):
        mapping = {}
        for ci, cell in enumerate(row):
            key = str(cell or '').strip().lower()
            if key in _SB_COLUMN_MAP:
                mapping[ci] = _SB_COLUMN_MAP[key]
        if 'invoice_number' in mapping.values():
            header_idx, col_map = i, mapping
            break
    if header_idx is None:
        result['errors'].append('Could not find the invoice columns. Please upload the Sleek Bill "Invoices Export" file.')
        return result

    for row in rows[header_idx + 1:]:
        if not any(c is not None and str(c).strip() != '' for c in row):
            continue
        data = {}
        for ci, field in col_map.items():
            if ci >= len(row):
                continue
            cell = row[ci]
            if field in _SB_DATE_FIELDS:
                data[field] = _parse_sleekbill_date(cell)
            elif field in _SB_AMOUNT_FIELDS:
                data[field] = _parse_sleekbill_amount(cell)
            else:
                data[field] = str(cell).strip() if cell is not None else ''
        inv_no = data.pop('invoice_number', '').strip()
        if not inv_no:
            result['skipped'] += 1
            continue
        data['source_file'] = source_file[:255]
        _, created = SleekBillInvoice.objects.update_or_create(invoice_number=inv_no, defaults=data)
        result['created' if created else 'updated'] += 1
    return result


class SleekBillInvoiceViewSet(viewsets.ModelViewSet):
    """Invoice Register — imported Sleek Bill invoices, mirroring the Sleek Bill
    list. Supports import, type/status/search/date filters, and summary totals."""
    serializer_class = SleekBillInvoiceSerializer
    pagination_class = ExpensePagination
    permission_classes = [IsAuthenticated, RequireSection(SECTION_SBINVOICE)]

    def get_queryset(self):
        qs = SleekBillInvoice.objects.all()
        inv_type = self.request.query_params.get('type')
        if inv_type:
            qs = qs.filter(invoice_type__iexact=inv_type)
        status_val = self.request.query_params.get('status')
        if status_val:
            qs = qs.filter(status__iexact=status_val)
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(Q(invoice_number__icontains=search) | Q(client_name__icontains=search) | Q(client_gstin__icontains=search))
        date_from = self.request.query_params.get('from')
        date_to = self.request.query_params.get('to')
        if date_from:
            qs = qs.filter(issue_date__gte=date_from)
        if date_to:
            qs = qs.filter(issue_date__lte=date_to)
        return qs.order_by('-issue_date', '-id')

    def list(self, request, *args, **kwargs):
        qs = self.filter_queryset(self.get_queryset())
        agg = qs.aggregate(
            amount=Coalesce(Sum('amount'), Decimal('0.00')),
            tax=Coalesce(Sum('tax'), Decimal('0.00')),
            total=Coalesce(Sum('total'), Decimal('0.00')),
            paid=Coalesce(Sum('amount_paid'), Decimal('0.00')),
            balance=Coalesce(Sum('balance'), Decimal('0.00')),
        )
        summary = {
            'count': qs.count(),
            'tax_invoice': qs.filter(invoice_type__iexact=SleekBillInvoice.TYPE_TAX).count(),
            'bill_of_supply': qs.filter(invoice_type__iexact=SleekBillInvoice.TYPE_BOS).count(),
            'amount': str(agg['amount']), 'tax': str(agg['tax']), 'total': str(agg['total']),
            'paid': str(agg['paid']), 'balance': str(agg['balance']),
        }
        page = self.paginate_queryset(qs)
        if page is not None:
            resp = self.get_paginated_response(self.get_serializer(page, many=True).data)
            resp.data['summary'] = summary
            return resp
        return Response({'results': self.get_serializer(qs, many=True).data, 'summary': summary})

    @action(detail=False, methods=['post'], url_path='import')
    def import_invoices(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'No file was uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            rows = _read_tabular(file_obj)
        except ValueError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        res = _import_sleekbill_rows(rows, file_obj.name)
        if res['created'] == 0 and res['updated'] == 0 and res['errors']:
            return Response({'detail': res['errors'][0], **res}, status=status.HTTP_400_BAD_REQUEST)
        parts = []
        if res['created']:
            parts.append(f"added {res['created']}")
        if res['updated']:
            parts.append(f"updated {res['updated']}")
        res['detail'] = 'Imported ' + (', '.join(parts) if parts else '0') + ' invoice(s).'
        return Response(res, status=status.HTTP_200_OK)

    @action(detail=False, methods=['delete'], url_path='clear')
    def clear(self, request):
        n, _ = SleekBillInvoice.objects.all().delete()
        return Response({'detail': f'Cleared {n} invoices.', 'deleted': n})

    @action(detail=True, methods=['get'], url_path='pdf')
    def pdf(self, request, pk=None):
        """Serve the attached Sleek Bill invoice PDF (fetched with auth by the UI)."""
        obj = self.get_object()
        if not obj.has_pdf:
            return Response({'detail': 'No PDF attached to this invoice.'}, status=status.HTTP_404_NOT_FOUND)
        resp = HttpResponse(bytes(obj.pdf_data), content_type='application/pdf')
        resp['Content-Disposition'] = f'inline; filename="{obj.invoice_number}.pdf"'
        return resp

    @action(detail=True, methods=['post'], url_path='upload-pdf')
    def upload_pdf(self, request, pk=None):
        """Attach a PDF to THIS specific invoice (reliable — no matching needed)."""
        obj = self.get_object()
        f = request.FILES.get('file')
        if not f:
            return Response({'detail': 'No file was uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
        obj.pdf_data = f.read()
        obj.pdf_name = f.name[:255]
        obj.save(update_fields=['pdf_data', 'pdf_name'])
        return Response({'detail': f'PDF attached to {obj.invoice_number}.'})

    @action(detail=False, methods=['post'], url_path='upload-pdfs')
    def upload_pdfs(self, request):
        """Bulk-attach PDFs, matching each file to an invoice by the invoice
        number found in its filename (e.g. 'RT26-27-SER-15.pdf')."""
        import re
        files = request.FILES.getlist('files') or request.FILES.getlist('file')
        if not files:
            return Response({'detail': 'No files were uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
        matched, unmatched = 0, []
        for f in files:
            m = re.search(r'RT\d{2}-\d{2}-(?:SER|REN)-\d+', f.name, re.IGNORECASE)
            inv = SleekBillInvoice.objects.filter(invoice_number__iexact=m.group(0)).first() if m else None
            if inv:
                inv.pdf_data = f.read()
                inv.pdf_name = f.name[:255]
                inv.save(update_fields=['pdf_data', 'pdf_name'])
                matched += 1
            else:
                unmatched.append(f.name)
        parts = [f'Attached {matched} PDF(s)']
        if unmatched:
            parts.append(f'{len(unmatched)} could not be matched by invoice number')
        return Response({'detail': ', '.join(parts) + '.', 'matched': matched, 'unmatched': unmatched})


class SubscriptionViewSet(viewsets.ModelViewSet):
    """CRUD for tracked service subscriptions (with renewal dates + reminders).
    Ordered by renewal date so the soonest-expiring appear first."""
    serializer_class = SubscriptionSerializer
    pagination_class = None
    permission_classes = [IsAuthenticated, RequireSection(SECTION_SUBSCRIPTION)]

    def get_queryset(self):
        qs = Subscription.objects.all()
        if self.request.query_params.get('active') == '1':
            qs = qs.filter(active=True)
        return qs.order_by('renewal_date', 'id')

